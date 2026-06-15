import openpyxl
from collections import defaultdict
from copy import deepcopy

# ======================== LOAD DATA ========================

def load_bank_04():
    wb = openpyxl.load_workbook(
        "/sessions/great-zealous-dirac/mnt/outputs/04 2026 Bank Recon Data_Bay City_161bct_MA.xlsx",
        read_only=True, data_only=True)
    ws = wb['DepositAccount_1680_050326_9519']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    bank = []
    for i, row in enumerate(rows[1:], start=2):
        date, amount, check_num, description = row
        if date is None or amount is None:
            continue
        bank.append({
            'row': i, 'source': 'Apr 2026',
            'date': str(date),
            'amount': float(amount),
            'check_number': str(check_num).strip() if check_num is not None else None,
            'description': str(description).strip() if description else ''
        })
    return bank

def load_gl_04():
    wb = openpyxl.load_workbook(
        "/sessions/great-zealous-dirac/mnt/outputs/04 2026 Bank Recon Data_Bay City_161bct_MA.xlsx",
        read_only=True, data_only=True)
    ws = wb['GL']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    gl = []
    for i, row in enumerate(rows[6:], start=7):
        if row[0] is None:
            continue
        date = row[2]
        if date is None:
            continue
        debit  = float(row[7] or 0)
        credit = float(row[8] or 0)
        if debit == 0 and credit == 0:
            continue
        gl.append({
            'row': i, 'source': 'Apr 2026',
            'date': str(date)[:10],
            'debit': debit, 'credit': credit,
            'amount': debit if debit > 0 else -credit,
            'reference': str(row[6]).strip() if row[6] is not None else None,
            'control': str(row[5]).strip() if row[5] is not None else '',
            'remarks': str(row[10]).strip() if row[10] is not None else '',
            'person_desc': str(row[4]).strip() if row[4] is not None else '',
            'description': str(row[4]).strip() if row[4] is not None else ''
        })
    return gl

def load_prev_unrec():
    wb = openpyxl.load_workbook(
        "/sessions/great-zealous-dirac/mnt/outputs/03 2026 Bay City Reconciliation.xlsx",
        read_only=True, data_only=True)
    ws = wb['Un-Reconcile transactions']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    prev_bank = []
    r = rows[2]
    if r[0] == 'Bay City' and r[6] is not None:
        prev_bank.append({
            'row': 3, 'source': 'Mar 2026 (prev)',
            'date': str(r[1]) if r[1] else '',
            'amount': float(r[6]),
            'check_number': str(r[3]).strip() if r[3] else None,
            'description': str(r[5]).strip() if r[5] else ''
        })
    prev_gl = []
    for i, r in enumerate(rows[7:94], start=8):
        if r[0] != 'Bay City' or r[6] is None:
            continue
        amt = float(r[6])
        prev_gl.append({
            'row': i, 'source': 'Mar 2026 (prev)',
            'date': str(r[1])[:10] if r[1] else '',
            'debit': amt if amt > 0 else 0.0,
            'credit': abs(amt) if amt < 0 else 0.0,
            'amount': amt,
            'reference': str(r[3]).strip() if r[3] else None,
            'control': str(r[2]).strip() if r[2] else '',
            'remarks': str(r[5]).strip() if r[5] else '',
            'person_desc': str(r[4]).strip() if r[4] else '',
            'description': str(r[4]).strip() if r[4] else ''
        })
    return prev_bank, prev_gl

bank_04 = load_bank_04()
gl_04   = load_gl_04()
prev_bank, prev_gl = load_prev_unrec()

all_bank = prev_bank + bank_04
all_gl   = prev_gl  + gl_04

# ======================== MATCHING ENGINE ========================

def amt_eq(a, b, tol=0.005):
    return abs(float(a) - float(b)) < tol

matched_bank = {}   # bank_idx -> {gl_indices, rule}
matched_gl   = set()
stagecoach_entries = []  # {bank_idx or None, gl_idx or None, type, amount, date, source}

# ---- Stagecoach sweep isolation ----
for bi, b in enumerate(all_bank):
    if 'STAGECOACH SWEEP' in b['description'].upper():
        stagecoach_entries.append({
            'type': 'Bank', 'row': b['row'], 'source': b['source'],
            'date': b['date'], 'amount': b['amount'],
            'description': b['description']
        })
        matched_bank[bi] = {'gl_indices': [], 'rule': 'STAGECOACH SWEEP (internal)'}

regular_bank = [bi for bi in range(len(all_bank)) if bi not in matched_bank]

def try_match(bank_idx, gl_candidates, rule, allow_multi=False):
    """Try to match a bank row to one or more GL rows. Returns True if matched."""
    if bank_idx in matched_bank:
        return False
    b = all_bank[bank_idx]
    bank_amt = b['amount']
    avail = [gi for gi in gl_candidates if gi not in matched_gl]

    if not allow_multi:
        if bank_amt >= 0:
            # Bank positive → GL debit
            for gi in avail:
                g = all_gl[gi]
                if g['debit'] > 0 and amt_eq(g['debit'], bank_amt):
                    matched_bank[bank_idx] = {'gl_indices': [gi], 'rule': rule}
                    matched_gl.add(gi)
                    return True
        else:
            # Bank negative → GL credit
            for gi in avail:
                g = all_gl[gi]
                if g['credit'] > 0 and amt_eq(g['credit'], abs(bank_amt)):
                    matched_bank[bank_idx] = {'gl_indices': [gi], 'rule': rule}
                    matched_gl.add(gi)
                    return True
    return False

# ---- RULE 1: Check number ----
for bi in regular_bank:
    b = all_bank[bi]
    if b['check_number']:
        chk = b['check_number']
        cands = [gi for gi, g in enumerate(all_gl)
                 if g['reference'] and g['reference'].strip() == chk]
        try_match(bi, cands, f'Check #{chk}')

# ---- RULE 2: PAYLEASE.COM → "PAYLEASE" in remarks ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'PAYLEASE.COM' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'PAYLEASE' in g['remarks'].upper()]
        try_match(bi, cands, 'PAYLEASE')

# ---- RULE 3: Bay City MHC LLC Settlement → "CKS" in remarks ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    du = b['description'].upper()
    if 'BAY CITY MHC LLC' in du and 'SETTLEMENT' in du:
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'CKS' in g['remarks'].upper()]
        try_match(bi, cands, 'CKS (Bay City MHC LLC Settlement)')

# ---- RULE 4: Bay City Settlement + Lakeshore Management → ACH/EFT Deposit ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    du = b['description'].upper()
    if 'BAY CITY SETTLEMENT' in du and 'LAKESHORE MANAGEMENT' in du:
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'ACH DEPOSIT' in g['remarks'].upper()
                    or 'EFT DEPOSIT' in g['remarks'].upper()]
        try_match(bi, cands, 'ACH/EFT Deposit (Bay City Settlement)')

# ---- RULE 5: BNKCD → "CC Deposit" in remarks ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'BNKCD' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'CC DEPOSIT' in g['remarks'].upper()]
        try_match(bi, cands, 'CC Deposit (BNKCD)')

# ---- RULE 6: INTELLIPAY BILLING → "Convenient Payments" ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'INTELLIPAY BILLING' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'CONVENIENT PAYMENTS' in g['remarks'].upper()]
        try_match(bi, cands, 'Convenient Payments (INTELLIPAY)')

# ---- RULE 7: LAKESHORE EMPLOYMENT → group GL by ref where person_desc has "LSE (v0000665)" ----
#   Multiple GL lines per payroll period sum to one bank wire
payroll_bank = [bi for bi in regular_bank
                if bi not in matched_bank
                and 'LAKESHORE EMPLOYMENT' in all_bank[bi]['description'].upper()]

lse_gl = [gi for gi, g in enumerate(all_gl)
          if 'LSE' in g['person_desc'].upper() and gi not in matched_gl]

# Group by reference (payroll batch)
lse_groups = defaultdict(list)
for gi in lse_gl:
    g = all_gl[gi]
    key = g['reference'] if g['reference'] else g['control']
    lse_groups[key].append(gi)

# Net credit per group
for ref_key, group_gi in lse_groups.items():
    avail_gi = [gi for gi in group_gi if gi not in matched_gl]
    if not avail_gi:
        continue
    net_credit = sum(all_gl[gi]['credit'] - all_gl[gi]['debit'] for gi in avail_gi)
    for bi in payroll_bank:
        if bi in matched_bank: continue
        if amt_eq(abs(all_bank[bi]['amount']), net_credit):
            matched_bank[bi] = {'gl_indices': avail_gi,
                                 'rule': f'LSE Payroll (LAKESHORE EMPLOYMENT) ref={ref_key}'}
            for gi in avail_gi:
                matched_gl.add(gi)
            break

# Also try matching individual LSE GL rows (benefits wires etc.)
lse_individual = [bi for bi in regular_bank
                  if bi not in matched_bank
                  and 'LAKESHORE EMPLOYMENT' in all_bank[bi]['description'].upper()]
for bi in lse_individual:
    b = all_bank[bi]
    cands = [gi for gi, g in enumerate(all_gl)
             if 'LSE' in g['person_desc'].upper() and gi not in matched_gl]
    try_match(bi, cands, 'LSE (LAKESHORE EMPLOYMENT) individual')

# ---- RULE 8: BILL PAY → "electric" in remarks ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'BILL PAY' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'ELECTRIC' in g['remarks'].upper()]
        try_match(bi, cands, 'Electric (BILL PAY)')

# ---- RULE 9: Return → "NSF RP" in remarks ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'RETURN' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'NSF RP' in g['remarks'].upper()
                    or 'NSF RP : ACH DEPOSIT' in g['remarks'].upper()]
        try_match(bi, cands, 'NSF Return')

# ---- FALLBACK: Amount-only matching for unmatched ----
# Build lookup tables keyed by rounded amount
remaining_bank = [bi for bi in regular_bank if bi not in matched_bank]
remaining_gl   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

bank_by_amt = defaultdict(list)
for bi in remaining_bank:
    bank_by_amt[round(all_bank[bi]['amount'], 2)].append(bi)

gl_by_amt = defaultdict(list)
for gi in remaining_gl:
    gl_by_amt[round(all_gl[gi]['amount'], 2)].append(gi)

# Unique-amount matches only
for amt, b_list in sorted(bank_by_amt.items()):
    if len(b_list) == 1 and amt in gl_by_amt and len(gl_by_amt[amt]) == 1:
        bi = b_list[0]
        gi = gl_by_amt[amt][0]
        if bi not in matched_bank and gi not in matched_gl:
            matched_bank[bi] = {'gl_indices': [gi], 'rule': 'Amount-only fallback'}
            matched_gl.add(gi)

# ======================== RESULTS ========================

unmatched_bank_idx = [bi for bi in regular_bank if bi not in matched_bank]
unmatched_gl_idx   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

total_bank_regular = len(regular_bank)
total_gl_all       = len(all_gl)
total_matched_bank = total_bank_regular - len(unmatched_bank_idx)
total_matched_gl   = total_gl_all - len(unmatched_gl_idx)

pct_bank = 100.0 * total_matched_bank / total_bank_regular if total_bank_regular else 0
pct_gl   = 100.0 * total_matched_gl   / total_gl_all       if total_gl_all else 0

print("=" * 60)
print("RECONCILIATION SUMMARY")
print("=" * 60)
print(f"Bank transactions (excl stagecoach) : {total_bank_regular}")
print(f"  Matched                           : {total_matched_bank}  ({pct_bank:.1f}%)")
print(f"  Unmatched                         : {len(unmatched_bank_idx)}")
print()
print(f"GL transactions                     : {total_gl_all}")
print(f"  Matched                           : {total_matched_gl}  ({pct_gl:.1f}%)")
print(f"  Unmatched                         : {len(unmatched_gl_idx)}")
print()
print(f"Stagecoach sweep entries (separate) : {len(stagecoach_entries)}")
print()

# Rule breakdown
rule_counts = defaultdict(int)
for bi, info in matched_bank.items():
    rule_counts[info['rule']] += 1
for rule, cnt in sorted(rule_counts.items(), key=lambda x: -x[1]):
    print(f"  {rule:50s} : {cnt} bank rows")

print()
print("=== UNMATCHED BANK TRANSACTIONS ===")
for bi in unmatched_bank_idx:
    b = all_bank[bi]
    print(f"  [{b['source']}] Row {b['row']} | {b['date']} | {b['amount']:>12,.2f} | {b['description'][:80]}")

print()
print("=== UNMATCHED GL TRANSACTIONS ===")
for gi in unmatched_gl_idx:
    g = all_gl[gi]
    print(f"  [{g['source']}] Row {g['row']} | {g['date']} | D:{g['debit']:>10,.2f} C:{g['credit']:>10,.2f} | rem={g['remarks'][:50]} | pdesc={g['person_desc'][:40]}")

# Save all_bank, all_gl, matched_bank, unmatched_bank_idx, unmatched_gl_idx,
# stagecoach_entries for the Excel output script
import pickle
with open('/sessions/great-zealous-dirac/mnt/outputs/recon_data.pkl', 'wb') as f:
    pickle.dump({
        'all_bank': all_bank,
        'all_gl': all_gl,
        'matched_bank': matched_bank,
        'unmatched_bank_idx': unmatched_bank_idx,
        'unmatched_gl_idx': unmatched_gl_idx,
        'stagecoach_entries': stagecoach_entries,
        'total_bank_regular': total_bank_regular,
        'total_matched_bank': total_matched_bank,
        'total_gl_all': total_gl_all,
        'total_matched_gl': total_matched_gl,
        'pct_bank': pct_bank,
        'pct_gl': pct_gl,
    }, f)
print("\nData saved for Excel generation.")


# ======================== RELOAD AND RE-RUN WITH EXTRA RULES ========================
# We append additional rules to catch more edge cases

import pickle
data = pickle.load(open('/sessions/great-zealous-dirac/mnt/outputs/recon_data.pkl','rb'))
all_bank    = data['all_bank']
all_gl      = data['all_gl']
matched_bank= data['matched_bank']
matched_gl  = data['matched_gl']
stagecoach_entries = data['stagecoach_entries']
regular_bank = [bi for bi in range(len(all_bank)) if bi not in matched_bank]

# ---- EXTRA RULE A: WT LAKESHORE MANAGEMENT → WC/Audit in GL remarks ----
# Bank wires to Lakeshore Management for insurance/audit → prev GL WC entries
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    du = b['description'].upper()
    if 'LAKESHORE MANAGEMENT' in du:
        # prefer prev-month entries first (were flagged to clear in April)
        for src_filter in ['Mar 2026 (prev)', 'Apr 2026']:
            cands = [gi for gi, g in enumerate(all_gl)
                     if gi not in matched_gl
                     and ('WC' in g['remarks'].upper() or 'AUDIT' in g['remarks'].upper())
                     and g['source'] == src_filter]
            if try_match(bi, cands, 'WC/Audit Insurance (WT Lakeshore Mgmt)'):
                break

# ---- EXTRA RULE B: DESKTOP CHECK DEPOSIT → GL with "Deposit" in remarks (excl accruals) ----
for bi in regular_bank:
    if bi in matched_bank: continue
    b = all_bank[bi]
    if 'DEPOSIT' in b['description'].upper():
        cands = [gi for gi, g in enumerate(all_gl)
                 if gi not in matched_gl
                 and g['debit'] > 0
                 and 'DEPOSIT' in g['remarks'].upper()
                 and 'ACCRUE' not in g['remarks'].upper()]
        try_match(bi, cands, 'Check Deposit')

# ---- EXTRA RULE C: Amount-only pass #2 allowing ties in bank if unique on GL side ----
remaining_bank2 = [bi for bi in regular_bank if bi not in matched_bank]
remaining_gl2   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

bank_by_amt2 = {}
for bi in remaining_bank2:
    amt = round(all_bank[bi]['amount'], 2)
    bank_by_amt2.setdefault(amt, []).append(bi)

gl_by_amt2 = {}
for gi in remaining_gl2:
    amt = round(all_gl[gi]['amount'], 2)
    gl_by_amt2.setdefault(amt, []).append(gi)

for amt, b_list in sorted(bank_by_amt2.items()):
    if len(b_list) == 1 and amt in gl_by_amt2 and len(gl_by_amt2[amt]) == 1:
        bi = b_list[0]
        gi = gl_by_amt2[amt][0]
        if bi not in matched_bank and gi not in matched_gl:
            matched_bank[bi] = {'gl_indices': [gi], 'rule': 'Amount-only fallback (pass 2)'}
            matched_gl.add(gi)

# ======================== FINAL STATS ========================
unmatched_bank_idx2 = [bi for bi in range(len(all_bank))
                       if bi not in matched_bank
                       and 'STAGECOACH SWEEP' not in all_bank[bi]['description'].upper()]
unmatched_gl_idx2   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

total_bank_regular2 = sum(1 for b in all_bank if 'STAGECOACH SWEEP' not in b['description'].upper())
total_matched_bank2 = total_bank_regular2 - len(unmatched_bank_idx2)
total_gl_all2       = len(all_gl)
total_matched_gl2   = total_gl_all2 - len(unmatched_gl_idx2)

pct_bank2 = 100.0 * total_matched_bank2 / total_bank_regular2
pct_gl2   = 100.0 * total_matched_gl2   / total_gl_all2

print("=" * 60)
print("FINAL RECONCILIATION SUMMARY (after extra rules)")
print("=" * 60)
print(f"Bank rows (excl stagecoach) : {total_bank_regular2}")
print(f"  Matched                   : {total_matched_bank2}  ({pct_bank2:.1f}%)")
print(f"  Unmatched                 : {len(unmatched_bank_idx2)}")
print(f"GL rows total               : {total_gl_all2}")
print(f"  Matched                   : {total_matched_gl2}  ({pct_gl2:.1f}%)")
print(f"  Unmatched                 : {len(unmatched_gl_idx2)}")
print(f"Stagecoach sweep (separate) : {len(stagecoach_entries)}")

# Save final data
with open('/sessions/great-zealous-dirac/mnt/outputs/recon_final.pkl', 'wb') as f:
    pickle.dump({
        'all_bank': all_bank,
        'all_gl': all_gl,
        'matched_bank': matched_bank,
        'matched_gl': matched_gl,
        'unmatched_bank_idx': unmatched_bank_idx2,
        'unmatched_gl_idx': unmatched_gl_idx2,
        'stagecoach_entries': stagecoach_entries,
        'stats': {
            'total_bank_regular': total_bank_regular2,
            'total_matched_bank': total_matched_bank2,
            'total_gl_all': total_gl_all2,
            'total_matched_gl': total_matched_gl2,
            'pct_bank': pct_bank2,
            'pct_gl': pct_gl2,
            'total_stagecoach': len(stagecoach_entries),
        }
    }, f)
print("Final data saved.")

