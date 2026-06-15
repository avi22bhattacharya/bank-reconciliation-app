"""
Build the Bay City bank reconciliation output workbook (new format, matching
"03 2026 Bay City Reconciliation.xlsx"):

  Sheets: Summary, Bank Statement, GL , Un-Reconcile transactions

This script:
  1. Loads the matched/unmatched bank & GL data produced by the
     reconciliation step (saved as recon_final.pkl).
  2. Pairs STAGECOACH SWEEP DEBIT/CREDIT bank transactions against each other
     (FIFO, exact-amount match) instead of giving them their own sheet.
  3. Assigns a Code to every bank and GL row:
       1  = unreconciled deposit (current month)
      -1  = unreconciled withdrawal (current month)
       2  = bank deposit matches GL deposit (both current month)
      -2  = bank withdrawal matches GL withdrawal (both current month)
       3  = cross-month deposit match (one side current, other side prior month)
      -3  = cross-month withdrawal match
       4  = Stagecoach Sweep credit leg of a matched pair
      -4  = Stagecoach Sweep debit leg of a matched pair
  4. Populates "Reconcile Date" on GL rows = date of the matched bank txn.
  5. Builds the 4-sheet workbook with the Summary formula structure
     (Bank Information / General Ledger Information, with
     Outstanding/Un-reconciled Deposits & Disbursements and a Variance
     check that should equal $0.00).

Run with: python3 build_excel.py
Requires recon_final.pkl (produced by the reconciliation matching step) in
the same directory as this script.

To reuse for a different month/property, update the CONFIG block below
(PREV_GL_COUNT, PROPERTY_*, PERIOD_*, BEGINNING_BALANCE, file paths).
"""
import pickle
from collections import Counter
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font

# ============================================================
# CONFIG - update these for each new month/property
# ============================================================
PKL_PATH = '/sessions/great-zealous-dirac/mnt/outputs/recon_final.pkl'
OUT_PATH = '/sessions/great-zealous-dirac/mnt/outputs/04_2026_Bay_City_Reconciliation_v2.xlsx'

# Number of GL rows (at the start of all_gl) that are carried over from the
# prior month's "Un-Reconcile" list (i.e. previously-outstanding GL items).
# Bank index 0 is similarly assumed to be the prior month's previously-
# unreconciled STAGECOACH SWEEP DEBIT.
PREV_GL_COUNT = 85

CUR = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
DATEFMT = 'mm/dd/yyyy'

PROPERTY_NAME = 'Bay City'
PROPERTY_TITLE = 'BAY CITY MHC, LLC'
PROPERTY_CODE = '161bct'
GL_ACCOUNT_LABEL = '11100161 BAY CITY MHC, LLC'
WFB_ACCOUNT = 'WFB Account number 4076381680'
PERIOD_START = date(2026, 4, 1)
DATE_RANGE_LABEL = 'Detail Date Range: 04/01/26 - 04/30/26  (Accural basis)'
BEGINNING_BALANCE = 344594.37


def parse_date(d):
    s = str(d)[:10]
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def amt_eq(a, b, tol=0.005):
    return abs(float(a) - float(b)) < tol


# ============================================================
# Load data
# ============================================================
data = pickle.load(open(PKL_PATH, 'rb'))
all_bank = data['all_bank']
all_gl = data['all_gl']
matched_bank = data['matched_bank']
matched_gl = data['matched_gl']

# ============================================================
# STEP 1: Stagecoach Sweep bank-to-bank pairing (FIFO, exact amount)
# ============================================================
sweep_debits, sweep_credits = [], []
for bi, b in enumerate(all_bank):
    du = b['description'].upper()
    if 'STAGECOACH SWEEP DEBIT' in du:
        sweep_debits.append(bi)
    elif 'STAGECOACH SWEEP CREDIT' in du:
        sweep_credits.append(bi)

sweep_debits.sort(key=lambda bi: parse_date(all_bank[bi]['date']))
sweep_credits.sort(key=lambda bi: parse_date(all_bank[bi]['date']))

sweep_pairs = []
used_credits = set()
unpaired_sweep = []
for d_bi in sweep_debits:
    d_amt = abs(all_bank[d_bi]['amount'])
    d_date = parse_date(all_bank[d_bi]['date'])
    match = None
    for c_bi in sweep_credits:
        if c_bi in used_credits:
            continue
        c_date = parse_date(all_bank[c_bi]['date'])
        if c_date is not None and c_date >= d_date and amt_eq(abs(all_bank[c_bi]['amount']), d_amt):
            match = c_bi
            break
    if match is not None:
        sweep_pairs.append((d_bi, match))
        used_credits.add(match)
    else:
        unpaired_sweep.append(d_bi)

unpaired_sweep += [c for c in sweep_credits if c not in used_credits]

print(f"Sweep pairs: {len(sweep_pairs)}, unpaired sweep entries: {len(unpaired_sweep)}")
for bi in unpaired_sweep:
    print(f"  UNPAIRED bank[{bi}]: {all_bank[bi]['date']} {all_bank[bi]['amount']} {all_bank[bi]['description']}")

# ============================================================
# STEP 2: Assign Code (1/-1/2/-2/3/-3/4/-4) and Reconcile Date
# ============================================================
bank_code = {}
gl_code = {}
gl_reconcile_date = {}

# 2a: sweep pairs -> -4 (debit leg) / 4 (credit leg)
for d_bi, c_bi in sweep_pairs:
    bank_code[d_bi] = -4
    bank_code[c_bi] = 4

# 2b: leftover unpaired sweep -> 1 / -1 by sign
for bi in unpaired_sweep:
    amt = all_bank[bi]['amount']
    bank_code[bi] = 1 if amt > 0 else -1

# 2c: regular matched_bank groups (skip the old empty-gl_indices stagecoach markers)
same_month_groups = []
cross_month_groups = []
for bi, info in matched_bank.items():
    gl_idxs = info['gl_indices']
    if not gl_idxs:
        continue  # old "STAGECOACH SWEEP (internal)" marker - handled in step 1
    b = all_bank[bi]
    is_deposit = b['amount'] > 0
    gl_months = set('prev' if gi < PREV_GL_COUNT else 'curr' for gi in gl_idxs)
    bank_month = 'prev' if bi == 0 else 'curr'
    if gl_months == {bank_month}:
        code = 2 if is_deposit else -2
        same_month_groups.append(bi)
    else:
        # cross-month (bank current / GL prior month, or vice versa)
        code = 3 if is_deposit else -3
        cross_month_groups.append(bi)
    bank_code[bi] = code
    for gi in gl_idxs:
        gl_code[gi] = code
        gl_reconcile_date[gi] = b['date']

print(f"Same-month matched groups (2/-2): {len(same_month_groups)}")
print(f"Cross-month matched groups (3/-3): {len(cross_month_groups)}")

# 2d: unmatched bank rows (not sweep, not matched) -> 1 / -1 by sign
unmatched_bank = []
for bi, b in enumerate(all_bank):
    if bi in bank_code:
        continue
    bank_code[bi] = 1 if b['amount'] > 0 else -1
    unmatched_bank.append(bi)

# 2e: unmatched GL rows -> 1 / -1 by sign
unmatched_gl = []
for gi, g in enumerate(all_gl):
    if gi in gl_code:
        continue
    gl_code[gi] = 1 if g['debit'] > 0 else -1
    unmatched_gl.append(gi)

assert len(bank_code) == len(all_bank), (len(bank_code), len(all_bank))
assert len(gl_code) == len(all_gl), (len(gl_code), len(all_gl))

print("Bank code distribution:", Counter(bank_code.values()))
print("GL code distribution:", Counter(gl_code.values()))

# ============================================================
# Build workbook
# ============================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ------------------------------------------------------------
# Sheet: Bank Statement
# ------------------------------------------------------------
ws = wb.create_sheet('Bank Statement')
ws.append(['Date', 'Amount', 'Check Number', 'Code', 'Description'])
for c in range(1, 6):
    ws.cell(row=1, column=c).font = Font(bold=True)

bank_order = sorted(range(len(all_bank)), key=lambda i: (parse_date(all_bank[i]['date']), i))

bank_row_of = {}
r = 2
for bi in bank_order:
    b = all_bank[bi]
    ws.cell(row=r, column=1, value=parse_date(b['date'])).number_format = DATEFMT
    cell = ws.cell(row=r, column=2, value=b['amount'])
    cell.number_format = CUR
    ws.cell(row=r, column=3, value=b['check_number'] if b['check_number'] else '')
    ws.cell(row=r, column=4, value=bank_code[bi])
    ws.cell(row=r, column=5, value=b['description'])
    bank_row_of[bi] = r
    r += 1

last_bank_row = r - 1
r += 1  # blank spacer
ws.cell(row=r, column=1, value=1)
ws.cell(row=r, column=2, value=f'=SUMIF(D2:D{last_bank_row},1,B2:B{last_bank_row})').number_format = CUR
ws.cell(row=r, column=3, value='Un-Reconciled Deposits')
r += 2
ws.cell(row=r, column=1, value=-1)
ws.cell(row=r, column=2, value=f'=SUMIF(D2:D{last_bank_row},-1,B2:B{last_bank_row})').number_format = CUR
ws.cell(row=r, column=3, value='Un-Reconciled Withdrawals')

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 90

print(f"Bank Statement: {last_bank_row - 1} data rows, totals at rows {r - 2} and {r}")

# ------------------------------------------------------------
# Sheet: GL  (trailing space, matches 03 2026 sheet name)
# ------------------------------------------------------------
wsg = wb.create_sheet('GL ')
wsg['B1'] = '                    General Ledger'
wsg['B2'] = f'Property: {PROPERTY_NAME} ({PROPERTY_CODE})'
wsg['B3'] = DATE_RANGE_LABEL
wsg['A4'] = f'{GL_ACCOUNT_LABEL}  (Bank Account)'
wsg.cell(row=4, column=9, value=BEGINNING_BALANCE).number_format = CUR

header = ['Property Name', 'Date', 'Control', 'Reference', 'Description', 'Remarks',
          'Debit', 'Credit', 'Balance', 'Reconcile Date', 'Code']
for c, h in enumerate(header, start=1):
    cell = wsg.cell(row=5, column=c, value=h)
    cell.font = Font(bold=True)

gl_row_of = {}
r = 6
running_balance = BEGINNING_BALANCE
for gi, g in enumerate(all_gl):
    is_prev = gi < PREV_GL_COUNT
    wsg.cell(row=r, column=1, value=PROPERTY_NAME)
    dt = parse_date(g['date'])
    wsg.cell(row=r, column=2, value=dt).number_format = DATEFMT
    wsg.cell(row=r, column=3, value=g['control'])
    wsg.cell(row=r, column=4, value=g['reference'])
    wsg.cell(row=r, column=5, value=g['person_desc'])
    wsg.cell(row=r, column=6, value=g['remarks'])
    debit = g['debit'] if g['debit'] else None
    credit = g['credit'] if g['credit'] else None
    if debit is not None:
        wsg.cell(row=r, column=7, value=debit).number_format = CUR
    if credit is not None:
        wsg.cell(row=r, column=8, value=credit).number_format = CUR
    if not is_prev:
        running_balance = running_balance + (g['debit'] or 0) - (g['credit'] or 0)
        wsg.cell(row=r, column=9, value=round(running_balance, 2)).number_format = CUR
    rd = gl_reconcile_date.get(gi)
    if rd:
        wsg.cell(row=r, column=10, value=parse_date(rd)).number_format = DATEFMT
    wsg.cell(row=r, column=11, value=gl_code[gi])
    gl_row_of[gi] = r
    r += 1

last_gl_row = r - 1
curr_first_row = gl_row_of[PREV_GL_COUNT]
curr_last_row = gl_row_of[len(all_gl) - 1]

# Totals row (current month only, current-month gl range)
totals_row = r
wsg.cell(row=totals_row, column=7, value=f'=SUM(G{curr_first_row}:G{curr_last_row})').number_format = CUR
wsg.cell(row=totals_row, column=8, value=f'=SUM(H{curr_first_row}:H{curr_last_row})').number_format = CUR
wsg.cell(row=totals_row, column=9, value=f'=I4+G{totals_row}-H{totals_row}').number_format = CUR

r = totals_row + 2
wsg.cell(row=r, column=6, value=1)
wsg.cell(row=r, column=7, value=f'=+SUMIF(K6:K{last_gl_row},1,G6:G{last_gl_row})+SUMIF(K6:K{last_gl_row},1,H6:H{last_gl_row})').number_format = CUR
wsg.cell(row=r, column=8, value='Un-Recorded deposits')
r += 2
wsg.cell(row=r, column=6, value=-1)
wsg.cell(row=r, column=7, value=f'=+SUMIF(K6:K{last_gl_row},-1,G6:G{last_gl_row})+SUMIF(K6:K{last_gl_row},-1,H6:H{last_gl_row})').number_format = CUR
wsg.cell(row=r, column=8, value='Un-Recorded withdrawal')

wsg.column_dimensions['A'].width = 11
wsg.column_dimensions['B'].width = 11
wsg.column_dimensions['C'].width = 12
wsg.column_dimensions['D'].width = 16
wsg.column_dimensions['E'].width = 45
wsg.column_dimensions['F'].width = 45
wsg.column_dimensions['G'].width = 13
wsg.column_dimensions['H'].width = 13
wsg.column_dimensions['I'].width = 13
wsg.column_dimensions['J'].width = 13
wsg.column_dimensions['K'].width = 6

GL_TOTALS_ROW = totals_row
print(f"GL: {last_gl_row - 5} data rows (rows 6-{last_gl_row}), totals row {totals_row}")

# ------------------------------------------------------------
# Sheet: Un-Reconcile transactions
# ------------------------------------------------------------
wsu = wb.create_sheet('Un-Reconcile transactions')

unmatched_bank = sorted([bi for bi, c in bank_code.items() if c in (1, -1)],
                         key=lambda bi: parse_date(all_bank[bi]['date']))
unmatched_gl_dep = sorted([gi for gi, c in gl_code.items() if c == 1],
                           key=lambda gi: parse_date(all_gl[gi]['date']))
unmatched_gl_wd = sorted([gi for gi, c in gl_code.items() if c == -1],
                          key=lambda gi: parse_date(all_gl[gi]['date']))

r = 1
wsu.cell(row=r, column=1, value='Bank Statement').font = Font(bold=True)
r += 1
bank_hdr = ['Property Name', 'Date', 'Control', 'Reference', 'Descriptions', 'Remarks', 'Amount', 'Comments']
for c, h in enumerate(bank_hdr, start=1):
    wsu.cell(row=r, column=c, value=h).font = Font(bold=True)
r += 1
bank_section_start = r
for bi in unmatched_bank:
    b = all_bank[bi]
    wsu.cell(row=r, column=1, value=PROPERTY_NAME)
    wsu.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
    wsu.cell(row=r, column=3, value=None)
    wsu.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
    wsu.cell(row=r, column=5, value=None)
    wsu.cell(row=r, column=6, value=b['description'])
    cell = wsu.cell(row=r, column=7, value=b['amount'])
    cell.number_format = CUR
    wsu.cell(row=r, column=8, value='AR' if b['amount'] > 0 else 'AP')
    r += 1
bank_section_end = r - 1
wsu.cell(row=r, column=7, value=f'=SUM(G{bank_section_start}:G{bank_section_end})').number_format = CUR
r += 2

wsu.cell(row=r, column=1, value='General Ledger').font = Font(bold=True)
r += 1
gl_hdr = ['Property Name', 'Date', 'Control', 'Reference', 'Description', 'Remarks', 'Amount', 'Comments']
for c, h in enumerate(gl_hdr, start=1):
    wsu.cell(row=r, column=c, value=h).font = Font(bold=True)
r += 1


def write_gl_section(rows_idx):
    global r
    start = r
    for gi in rows_idx:
        g = all_gl[gi]
        wsu.cell(row=r, column=1, value=PROPERTY_NAME)
        wsu.cell(row=r, column=2, value=parse_date(g['date'])).number_format = DATEFMT
        wsu.cell(row=r, column=3, value=g['control'])
        wsu.cell(row=r, column=4, value=g['reference'])
        wsu.cell(row=r, column=5, value=g['person_desc'])
        wsu.cell(row=r, column=6, value=g['remarks'])
        amount = g['debit'] if g['debit'] else -g['credit']
        cell = wsu.cell(row=r, column=7, value=amount)
        cell.number_format = CUR
        wsu.cell(row=r, column=8, value='AR' if amount > 0 else 'AP')
        r += 1
    end = r - 1
    if end >= start:
        wsu.cell(row=r, column=7, value=f'=SUM(G{start}:G{end})').number_format = CUR
        r += 1
    return start, end


dep_start, dep_end = write_gl_section(unmatched_gl_dep)
r += 1
wd_start, wd_end = write_gl_section(unmatched_gl_wd)

wsu.column_dimensions['A'].width = 11
wsu.column_dimensions['B'].width = 11
wsu.column_dimensions['C'].width = 14
wsu.column_dimensions['D'].width = 18
wsu.column_dimensions['E'].width = 50
wsu.column_dimensions['F'].width = 50
wsu.column_dimensions['G'].width = 13
wsu.column_dimensions['H'].width = 9

print(f"Un-Reconcile transactions: bank rows {bank_section_start}-{bank_section_end}, "
      f"GL deposits {dep_start}-{dep_end}, GL withdrawals {wd_start}-{wd_end}")

# ------------------------------------------------------------
# Sheet: Summary
# ------------------------------------------------------------
wss = wb.create_sheet('Summary', 0)  # make it first sheet

wss['B1'] = ' '
wss['B2'] = f'Bank Reconciliation - {PROPERTY_TITLE}'
wss['B3'] = WFB_ACCOUNT
wss['B4'] = PERIOD_START
wss['B4'].number_format = DATEFMT
wss['B2'].font = Font(bold=True)
wss['B5'] = 'Bank Information'
wss['B5'].font = Font(bold=True)

r = 7
wss.cell(row=r, column=6, value='Balance per bank at the month end')
wss.cell(row=r, column=8, value=0).number_format = CUR
H7_row = r

r += 2
wss.cell(row=r, column=2, value='Outstanding Bank Deposits:').font = Font(italic=True)
r += 1
items_first_row = r
ub_dep = sorted([bi for bi, c in bank_code.items() if c == 1], key=lambda bi: parse_date(all_bank[bi]['date']))
ub_wd = sorted([bi for bi, c in bank_code.items() if c == -1], key=lambda bi: parse_date(all_bank[bi]['date']))

for bi in ub_dep:
    b = all_bank[bi]
    wss.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
    wss.cell(row=r, column=3, value='AR')
    wss.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
    wss.cell(row=r, column=5, value=b['description'])
    wss.cell(row=r, column=7, value=b['amount']).number_format = CUR
    r += 1

r += 1
wss.cell(row=r, column=2, value='Outstanding Bank Disbursement:').font = Font(italic=True)
r += 1
for bi in ub_wd:
    b = all_bank[bi]
    wss.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
    wss.cell(row=r, column=3, value='AP')
    wss.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
    wss.cell(row=r, column=5, value=b['description'])
    wss.cell(row=r, column=7, value=b['amount']).number_format = CUR
    r += 1

items_last_row = r - 1
r += 1
wss.cell(row=r, column=6, value='Outstanding Items Total')
wss.cell(row=r, column=8, value=f'=+SUM(G{items_first_row}:G{items_last_row})').number_format = CUR
H16_row = r

r += 2
wss.cell(row=r, column=6, value='Adjusted bank balance')
wss.cell(row=r, column=8, value=f'=+H{H7_row}-H{H16_row}').number_format = CUR
H18_row = r

r += 2
wss.cell(row=r, column=2, value=f'General Ledger Information - {GL_ACCOUNT_LABEL}').font = Font(bold=True)

r += 2
wss.cell(row=r, column=6, value=' Balance per TB at the month end')
wss.cell(row=r, column=8, value=f"=+'GL '!I{GL_TOTALS_ROW}").number_format = CUR
H22_row = r

r += 2
wss.cell(row=r, column=2, value='Un-reconcile Outstanding Deposits:').font = Font(italic=True)
r += 1
outstanding_dep_first_row = r
gl_dep_outstanding = sorted([gi for gi, c in gl_code.items() if c == 1 and gi < PREV_GL_COUNT],
                             key=lambda gi: parse_date(all_gl[gi]['date']))
gl_dep_curr = sorted([gi for gi, c in gl_code.items() if c == 1 and gi >= PREV_GL_COUNT],
                      key=lambda gi: parse_date(all_gl[gi]['date']))
gl_wd_outstanding = sorted([gi for gi, c in gl_code.items() if c == -1 and gi < PREV_GL_COUNT],
                            key=lambda gi: parse_date(all_gl[gi]['date']))
gl_wd_curr = sorted([gi for gi, c in gl_code.items() if c == -1 and gi >= PREV_GL_COUNT],
                     key=lambda gi: parse_date(all_gl[gi]['date']))


def write_summary_gl_items(rows_idx):
    global r
    for gi in rows_idx:
        g = all_gl[gi]
        amount = g['debit'] if g['debit'] else -g['credit']
        wss.cell(row=r, column=2, value=parse_date(g['date'])).number_format = DATEFMT
        wss.cell(row=r, column=3, value='AR' if amount > 0 else 'AP')
        wss.cell(row=r, column=4, value=g['control'])
        wss.cell(row=r, column=5, value=g['remarks'])
        wss.cell(row=r, column=7, value=amount).number_format = CUR
        r += 1


write_summary_gl_items(gl_dep_outstanding)

r += 1
wss.cell(row=r, column=2, value="Un-reconcile Deposits: Apr '26").font = Font(italic=True)
r += 1
dep_first_row = r
write_summary_gl_items(gl_dep_curr)
dep_last_row = r - 1
r += 1
wss.cell(row=r, column=6, value='Total ')
G37_row = r
wss.cell(row=r, column=7, value=f'=SUM(G{outstanding_dep_first_row}:G{dep_last_row})').number_format = CUR

r += 2
wss.cell(row=r, column=2, value='Un-reconciled Outstanding Disbursement:').font = Font(italic=True)
r += 1
wd_outstanding_first_row = r
write_summary_gl_items(gl_wd_outstanding)

r += 1
wss.cell(row=r, column=2, value="Un-reconcile Disbursement: Apr '26").font = Font(italic=True)
r += 1
write_summary_gl_items(gl_wd_curr)
wd_last_row = r - 1
r += 1
wss.cell(row=r, column=6, value='Total ')
G59_row = r
wss.cell(row=r, column=7, value=f'=SUM(G{wd_outstanding_first_row}:G{wd_last_row})').number_format = CUR

r += 2
wss.cell(row=r, column=6, value='Sub-Total ')
G61_row = r
wss.cell(row=r, column=7, value=f'=+G{G59_row}+G{G37_row}').number_format = CUR

r += 2
wss.cell(row=r, column=6, value='Adjusted book balance')
wss.cell(row=r, column=8, value=f'=+H{H22_row}-G{G61_row}').number_format = CUR
H63_row = r

r += 2
wss.cell(row=r, column=6, value='Variance (s/b zero)')
wss.cell(row=r, column=8, value=f'=+H{H18_row}-H{H63_row}').number_format = CUR

wss.column_dimensions['B'].width = 12
wss.column_dimensions['C'].width = 6
wss.column_dimensions['D'].width = 16
wss.column_dimensions['E'].width = 55
wss.column_dimensions['F'].width = 32
wss.column_dimensions['G'].width = 14
wss.column_dimensions['H'].width = 14

print(f"Summary sheet built, {r} rows total")

wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}")
