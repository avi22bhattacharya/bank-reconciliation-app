"""Prior unreconciled items: DB fetch, bootstrap parsing from Excel,
and per-engine injection adapters.

Canonical item shapes (floats, ISO dates):
  bank: {date, amount, check_number, description}
  gl:   {date, control, reference, description, remarks, debit, credit,
         deposit_number, tenant_code, property_label}
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook, load_workbook

from core import db


@dataclass
class PriorItems:
    bank: list[dict] = field(default_factory=list)
    gl: list[dict] = field(default_factory=list)

    def __bool__(self):
        return bool(self.bank or self.gl)


def clean_check(v) -> str | None:
    """Normalize check numbers: 1903.0 / '1903.0' / ' 1903' → '1903'."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s or None


def _as_date(v) -> str:
    return db.iso_date(v)


# ── DB fetch / seed ──────────────────────────────────────────────────────────

def fetch(con, property_code: str) -> PriorItems:
    """All still-unmatched transactions for the property, in stable order.

    Each item carries its txn_hash so the post-run persist step can update
    the existing row directly instead of recomputing a hash (recomputation
    would mis-handle content-identical duplicates)."""
    prior = PriorItems()
    for r in db.unmatched_bank(con, property_code):
        prior.bank.append({
            "txn_hash": r["txn_hash"],
            "date": r["date"], "amount": r["amount_cents"] / 100.0,
            "check_number": r["check_number"], "description": r["description"] or "",
        })
    for r in db.unmatched_gl(con, property_code):
        prior.gl.append({
            "txn_hash": r["txn_hash"],
            "date": r["date"], "control": r["control"] or "",
            "reference": r["reference"] or "", "description": r["description"] or "",
            "remarks": r["remarks"] or "",
            "debit": r["debit_cents"] / 100.0, "credit": r["credit_cents"] / 100.0,
            "deposit_number": r["deposit_number"] or "",
            "tenant_code": r["tenant_code"] or "",
            "property_label": r["property_label"] or "",
        })
    return prior


def seed_db(con, property_code: str, prior: PriorItems, source_period: str) -> int:
    """Insert bootstrap items as unmatched rows. Content-identical rows get
    occurrence-indexed hashes so none collapse. Returns number inserted."""
    from collections import Counter
    n = 0
    occ = Counter()
    for b in prior.bank:
        cents = db.to_cents(b["amount"])
        chk = clean_check(b["check_number"])
        key = db.bank_content_key(property_code, b["date"], cents, chk, b["description"])
        h = db.bank_hash(property_code, source_period, occ[key], b["date"], cents,
                         chk, b["description"])
        occ[key] += 1
        cur = con.execute("""
            INSERT OR IGNORE INTO bank_txns
              (txn_hash, property_code, date, amount_cents, check_number,
               description, status, source_period)
            VALUES (?,?,?,?,?,?,'unmatched',?)
        """, (h, property_code, _as_date(b["date"]), cents, chk,
              b["description"], source_period))
        n += cur.rowcount
    occ = Counter()
    for g in prior.gl:
        dc, cc = db.to_cents(g["debit"]), db.to_cents(g["credit"])
        key = db.gl_content_key(property_code, g["date"], g["control"], g["reference"],
                                dc, cc, g["description"], g["remarks"])
        h = db.gl_hash(property_code, source_period, occ[key], g["date"], g["control"],
                       g["reference"], dc, cc, g["description"], g["remarks"])
        occ[key] += 1
        cur = con.execute("""
            INSERT OR IGNORE INTO gl_txns
              (txn_hash, property_code, property_label, date, control, reference,
               description, remarks, debit_cents, credit_cents, deposit_number,
               tenant_code, status, source_period)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'unmatched',?)
        """, (h, property_code, g.get("property_label", ""), _as_date(g["date"]),
              g["control"], g["reference"], g["description"], g["remarks"],
              dc, cc, g.get("deposit_number", ""), g.get("tenant_code", ""),
              source_period))
        n += cur.rowcount
    con.commit()
    return n


# ── Bootstrap parsing ────────────────────────────────────────────────────────

def _header_map(row) -> dict[str, int]:
    """Header name → column index; first occurrence wins (Yardi sheets have
    two 'Description' columns and the first is the canonical one)."""
    h = {}
    for i, v in enumerate(row):
        if v in (None, ""):
            continue
        h.setdefault(str(v).strip().lower(), i)
    return h


def parse_open_items(path) -> PriorItems:
    """Parse an "Open Items" workbook (sheets: Un-Reconcile Bank / Un-Reconcile GL).

    Column positions differ between Yardi and MA exports, so columns are
    resolved by header name.
    """
    wb = load_workbook(path, data_only=True)
    prior = PriorItems()

    if "Un-Reconcile Bank" in wb.sheetnames:
        rows = list(wb["Un-Reconcile Bank"].iter_rows(values_only=True))
        hdr_i = next((i for i, r in enumerate(rows)
                      if r and str(r[0]).strip().lower() == "date"), None)
        if hdr_i is not None:
            h = _header_map(rows[hdr_i])
            desc_i = h.get("description", h.get("descriptions"))
            for r in rows[hdr_i + 1:]:
                if r[h["date"]] is None or r[h["amount"]] is None:
                    continue
                prior.bank.append({
                    "date": _as_date(r[h["date"]]),
                    "amount": float(r[h["amount"]]),
                    "check_number": clean_check(r[h.get("check number", 2)]),
                    "description": str(r[desc_i]).strip() if desc_i is not None and r[desc_i] else "",
                })

    if "Un-Reconcile GL" in wb.sheetnames:
        rows = list(wb["Un-Reconcile GL"].iter_rows(values_only=True))
        h = _header_map(rows[0])
        for r in rows[1:]:
            amt = r[h["amount"]] if h.get("amount") is not None else None
            if amt is None or not isinstance(amt, (int, float)):
                continue
            ctl = str(r[h["control"]]).strip() if h.get("control") is not None and r[h["control"]] else ""
            ref = str(r[h["reference"]]).strip() if h.get("reference") is not None and r[h["reference"]] else ""
            desc = str(r[h["description"]]).strip() if h.get("description") is not None and r[h["description"]] else ""
            rem = str(r[h["remarks"]]).strip() if h.get("remarks") is not None and r[h["remarks"]] else ""
            if not (ctl or ref or desc):
                continue
            amt = float(amt)
            prior.gl.append({
                "date": _as_date(r[h["date"]]), "control": ctl, "reference": ref,
                "description": desc, "remarks": rem,
                "debit": round(amt, 2) if amt > 0 else 0.0,
                "credit": round(-amt, 2) if amt < 0 else 0.0,
                "deposit_number": "", "tenant_code": "",
                "property_label": str(r[h["property name"]]).strip()
                                  if h.get("property name") is not None and r[h["property name"]] else "",
            })
    wb.close()
    return prior


def parse_prior_output(path) -> PriorItems:
    """Parse a prior reconciliation output's "Un-Reconcile transactions" sheet.

    Sections are located dynamically by their header labels ("Bank Statement"/
    "Bank" then "General Ledger"/"GL") — no hardcoded row ranges. Total rows
    (blank Property Name) are skipped. In this layout the bank description
    lives under the "Remarks" column and the check number under "Reference".
    """
    wb = load_workbook(path, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == "un-reconcile transactions"), None)
    if sheet is None:
        wb.close()
        return PriorItems()
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    prior = PriorItems()
    section = None  # None | 'bank' | 'gl'
    for r in rows:
        c0 = str(r[0]).strip() if r[0] is not None else ""
        c0l = c0.lower()
        if c0l in ("bank statement", "bank"):
            section = "bank"
            continue
        if c0l in ("general ledger", "gl"):
            section = "gl"
            continue
        if c0l == "property name":   # column header row
            continue
        if section is None or not c0:  # totals / spacers have a blank first col
            continue
        if len(r) < 7 or r[6] is None or not isinstance(r[6], (int, float)):
            continue
        amt = float(r[6])
        if section == "bank":
            prior.bank.append({
                "date": _as_date(r[1]), "amount": amt,
                "check_number": clean_check(r[3]),
                "description": str(r[5]).strip() if r[5] else "",
            })
        else:
            prior.gl.append({
                "date": _as_date(r[1]),
                "control": str(r[2]).strip() if r[2] else "",
                "reference": str(r[3]).strip() if r[3] else "",
                "description": str(r[4]).strip() if r[4] else "",
                "remarks": str(r[5]).strip() if r[5] else "",
                "debit": round(amt, 2) if amt > 0 else 0.0,
                "credit": round(-amt, 2) if amt < 0 else 0.0,
                "deposit_number": "", "tenant_code": "",
                "property_label": c0,
            })
    return prior


def parse_bootstrap(path) -> PriorItems:
    """Dispatch on workbook shape: Open Items file or prior output workbook."""
    wb = load_workbook(path, read_only=True)
    names = [s.strip().lower() for s in wb.sheetnames]
    wb.close()
    if "un-reconcile gl" in names or "un-reconcile bank" in names:
        return parse_open_items(path)
    if "un-reconcile transactions" in names:
        return parse_prior_output(path)
    raise ValueError(
        "Unrecognized prior open-items workbook: expected sheets "
        "'Un-Reconcile Bank'/'Un-Reconcile GL' or 'Un-Reconcile transactions'")


# ── Engine injection adapters ────────────────────────────────────────────────

def to_ma_prev(prior: PriorItems, prev_label: str) -> tuple[list[dict], list[dict]]:
    """PriorItems → the (prev_bank, prev_gl) lists reconcile_full.run expects."""
    prev_bank = []
    for i, b in enumerate(prior.bank, start=1):
        prev_bank.append({
            "row": i, "source": prev_label, "date": b["date"],
            "amount": float(b["amount"]),
            "check_number": b["check_number"],
            "description": b["description"],
        })
    prev_gl = []
    for i, g in enumerate(prior.gl, start=1):
        amt = round(g["debit"] - g["credit"], 2)
        prev_gl.append({
            "row": i, "source": prev_label, "date": g["date"],
            "debit": g["debit"], "credit": g["credit"], "amount": amt,
            "reference": g["reference"] or None, "control": g["control"],
            "remarks": g["remarks"], "person_desc": g["description"],
            "description": g["description"],
        })
    return prev_bank, prev_gl


YARDI_PREV_HEADER = ["Property", "Property Name", "Date", "Period", "Description",
                     "Control", "Reference", "Description", "Amount", "Remarks",
                     "Comment"]


def write_yardi_prev_workbook(prior: PriorItems, out_path, property_code: str):
    """Write prior GL items as an 11-column "Un-Reconcile GL" sheet so they
    flow through mh_recon (deposit-number columns land at indices 11-13,
    matching what reconcile_ph reads)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Un-Reconcile GL"
    ws.append(YARDI_PREV_HEADER)
    for g in prior.gl:
        d = datetime.fromisoformat(g["date"]) if g["date"] else None
        amt = round(g["debit"] - g["credit"], 2)
        ws.append([property_code, g["property_label"], d, None, g["description"],
                   g["control"], g["reference"], g["description"], amt,
                   g["remarks"], "AR" if amt > 0 else "AP"])
    wb.save(out_path)
    return out_path


def append_prev_bank_rows(bank_xlsx, bank_sheet: str, prior: PriorItems) -> int:
    """Append prior unreconciled bank rows to the converted bank statement
    sheet (4-col layout) so reconcile_ph sees them as part of the data.
    Returns the sheet row of the first appended item (engine ids are
    BANK-<sheet row>, which lets the caller map them back to prior items)."""
    wb = load_workbook(bank_xlsx)
    ws = wb[bank_sheet]
    first_row = ws.max_row + 1
    if not prior.bank:
        wb.close()
        return first_row
    for b in prior.bank:
        chk = b["check_number"]
        if chk and chk.isdigit():
            chk = int(chk)
        d = b["date"]
        if d:
            d = datetime.fromisoformat(d).strftime("%m/%d/%Y")
        ws.append([d, b["amount"], chk, b["description"]])
    wb.save(bank_xlsx)
    return first_row


def patch_deposit_numbers(enriched_path, prior: PriorItems,
                          sheet: str = "Un-Reconcile GL"):
    """Restore persisted deposit numbers onto the mh_recon-enriched prev sheet.

    Prior items usually predate the current month's Deposit Register, so
    re-enrichment leaves their Deposit Number blank; without this patch the
    P6/P7 deposit-grouping passes silently skip carryover items. Rows are
    matched by order (mh_recon preserves input row order; data starts at
    row 2). Only non-empty stored values overwrite."""
    if not prior.gl:
        return
    wb = load_workbook(enriched_path)
    ws = wb[sheet]
    dep_col = len(YARDI_PREV_HEADER) + 3   # 11 source cols + Tenant/UID/Deposit
    for i, g in enumerate(prior.gl):
        if g.get("deposit_number"):
            ws.cell(row=2 + i, column=dep_col, value=g["deposit_number"])
    wb.save(enriched_path)
