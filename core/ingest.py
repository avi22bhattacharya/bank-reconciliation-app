"""Upload ingestion: .xls→.xlsx conversion, sheet auto-detection,
property/period detection from the GL metadata rows."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

BANK_HEADER = ["date", "amount", "check number", "description"]
GL_HEADER_PREFIX = ["property", "property name", "date", "period"]


def convert_to_xlsx(path: str | Path, out_path: str | Path | None = None) -> Path:
    """Convert a workbook to .xlsx (no-op copy of path if already .xlsx).

    Reads via pandas (xlrd handles legacy .xls) with header=None so every
    sheet round-trips cell-for-cell, preserving datetimes and numbers.
    Replaces the LibreOffice dependency of the original scripts.
    """
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        return path
    out_path = Path(out_path) if out_path else path.with_suffix(".xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in pd.read_excel(path, sheet_name=None, header=None).items():
        # openpyxl caps sheet names at 31 chars (xlrd may report longer)
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        for row in df.itertuples(index=False):
            ws.append([None if pd.isna(v) else v for v in row])
    wb.save(out_path)
    return out_path


def sheet_names(path: str | Path) -> list[str]:
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    return pd.ExcelFile(path).sheet_names


def _first_row(path: str | Path, sheet: str) -> list:
    df = pd.read_excel(path, sheet_name=sheet, header=None, nrows=1)
    return [] if df.empty else df.iloc[0].tolist()


def detect_bank_sheet(path: str | Path) -> str | None:
    names = sheet_names(path)
    for n in names:
        if re.match(r"^DepositAccount", str(n)):
            return n
    for n in names:
        row = [str(v).strip().lower() for v in _first_row(path, n)[:4]]
        if row == BANK_HEADER:
            return n
    return None


def detect_gl_sheet(path: str | Path) -> str | None:
    names = sheet_names(path)
    if "GL" in names:
        return "GL"
    for n in names:
        df = pd.read_excel(path, sheet_name=n, header=None, nrows=7)
        for i in range(len(df)):
            row = [str(v).strip().lower() for v in df.iloc[i, :4].tolist()]
            if row == GL_HEADER_PREFIX:
                return n
    return None


@dataclass
class GLMetadata:
    property_code: str = ""      # canonical key: codes joined by '^' (e.g. 113phv^114epr)
    property_name: str = ""
    gl_type_guess: str = ""      # 'yardi' | 'ma' | ''
    period: str = ""             # YYYY-MM
    gl_account_number: str = ""  # e.g. 11100113
    raw_rows: list = field(default_factory=list)


def detect_gl_metadata(path: str | Path, gl_sheet: str = "GL") -> GLMetadata:
    """Parse the GL metadata block (rows 0-6, col 0).

    Yardi:  'Property =  113phv 114epr'
    MA:     'Bay City (161bct)'
    Both:   'Period = Mar 2026', account number in the row after the header.
    """
    df = pd.read_excel(path, sheet_name=gl_sheet, header=None, nrows=7)
    cells = ["" if pd.isna(v) else str(v).strip() for v in df.iloc[:, 0].tolist()]
    meta = GLMetadata(raw_rows=cells)

    for cell in cells:
        m = re.match(r"^Property\s*=\s*(.+)$", cell)
        if m:
            codes = m.group(1).split()
            meta.property_code = "^".join(codes)
            meta.gl_type_guess = "yardi"
            break
        m = re.match(r"^(.+?)\s*\((\w+)\)\s*$", cell)
        if m and cell.lower() != "general ledger":
            meta.property_name = m.group(1).strip()
            meta.property_code = m.group(2)
            meta.gl_type_guess = "ma"
            break

    for cell in cells:
        m = re.search(r"Period\s*=\s*([A-Za-z]{3,9})\s+(\d{4})", cell)
        if m:
            try:
                month = pd.to_datetime(f"{m.group(1)} {m.group(2)}", format="%b %Y")
            except ValueError:
                month = pd.to_datetime(f"{m.group(1)} {m.group(2)}")
            meta.period = month.strftime("%Y-%m")
            break

    for cell in cells:
        if re.fullmatch(r"\d{6,10}", cell):
            meta.gl_account_number = cell
            break

    return meta


def period_label(period: str) -> str:
    """'2026-03' → 'Mar 2026' (the label style the engines embed in sources)."""
    return pd.to_datetime(period + "-01").strftime("%b %Y")
