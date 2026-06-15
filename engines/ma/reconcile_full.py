"""
Manage America reconciliation engine (refactored from
"Manage America Bank recon/reconcile_full.py"; matching rules unchanged).

Changes vs the original:
  - run(workbook_path, bank_sheet, ...) replaces hardcoded /sessions/ paths.
  - Prior unreconciled items are passed in as prev_bank/prev_gl lists of dicts
    (same shape the old load_prev_unrec() produced) instead of being parsed
    from the prior output workbook with hardcoded row ranges.
  - The mid-script pickle dump/reload is removed: the original dumped a dict
    WITHOUT 'matched_gl' and then reloaded data['matched_gl'] (a guaranteed
    KeyError on a fresh run). The extra rules now continue inline with
    matched_bank/matched_gl still in scope.
  - Returns the final results dict (recon_final.pkl payload shape, plus
    n_prev_bank/n_prev_gl for the output writer); optionally pickles it.
"""
import pickle
from collections import defaultdict

import openpyxl

# ======================== LOAD DATA ========================

def load_bank(workbook_path, bank_sheet, source_label):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[bank_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    bank = []
    for i, row in enumerate(rows[1:], start=2):
        date, amount, check_num, description = row[:4]
        if date is None or amount is None:
            continue
        bank.append({
            'row': i, 'source': source_label,
            'date': str(date),
            'amount': float(amount),
            'check_number': str(check_num).strip() if check_num is not None else None,
            'description': str(description).strip() if description else ''
        })
    return bank


def load_gl(workbook_path, gl_sheet, source_label):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[gl_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    gl = []
    for i, row in enumerate(rows[6:], start=7):
        if row[0] is None:
            continue
        date = row[2]
        if date is None:
            continue
        debit  = float(row[7] or 0)
        credit = float(row[8] or 0)
        if debit == 0 and credit == 0:
            continue
        gl.append({
            'row': i, 'source': source_label,
            'date': str(date)[:10],
            'debit': debit, 'credit': credit,
            'amount': debit if debit > 0 else -credit,
            'reference': str(row[6]).strip() if row[6] is not None else None,
            'control': str(row[5]).strip() if row[5] is not None else '',
            'remarks': str(row[10]).strip() if row[10] is not None else '',
            'person_desc': str(row[4]).strip() if row[4] is not None else '',
            'description': str(row[4]).strip() if row[4] is not None else ''
        })
    return gl


def read_gl_opening_balance(workbook_path, gl_sheet):
    """GL opening balance: row 7, col J (index 9) — the balance row above the data."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[gl_sheet]
    row = next(ws.iter_rows(min_row=7, max_row=7, values_only=True), None)
    wb.close()
    if row and row[9] is not None:
        return float(row[9])
    return 0.0


# ======================== ENTRY POINT ========================

def run(workbook_path, bank_sheet, gl_sheet="GL",
        prev_bank=(), prev_gl=(), current_label="current",
        bank_workbook_path=None, pkl_out=None) -> dict:
    """Run the full MA reconciliation and return the results dict.

    prev_bank/prev_gl: lists of dicts in the loader shape above, with a
    'source' containing "prev". bank_workbook_path lets the bank sheet live
    in a different workbook than the GL (defaults to workbook_path).
    """
    bank_cur = load_bank(bank_workbook_path or workbook_path, bank_sheet, current_label)
    gl_cur   = load_gl(workbook_path, gl_sheet, current_label)
    prev_bank = list(prev_bank)
    prev_gl   = list(prev_gl)

    all_bank = prev_bank + bank_cur
    all_gl   = prev_gl  + gl_cur

    # ======================== MATCHING ENGINE ========================

    def amt_eq(a, b, tol=0.005):
        return abs(float(a) - float(b)) < tol

    matched_bank = {}   # bank_idx -> {gl_indices, rule}
    matched_gl   = set()
    stagecoach_entries = []

    # ---- Stagecoach sweep isolation ----
    for bi, b in enumerate(all_bank):
        if 'STAGECOACH SWEEP' in b['description'].upper():
            stagecoach_entries.append({
                'type': 'Bank', 'row': b['row'], 'source': b['source'],
                'date': b['date'], 'amount': b['amount'],
                'description': b['description']
            })
            matched_bank[bi] = {'gl_indices': [], 'rule': 'STAGECOACH SWEEP (internal)'}

    regular_bank = [bi for bi in range(len(all_bank)) if bi not in matched_bank]

    def try_match(bank_idx, gl_candidates, rule, allow_multi=False):
        """Try to match a bank row to one or more GL rows. Returns True if matched."""
        if bank_idx in matched_bank:
            return False
        b = all_bank[bank_idx]
        bank_amt = b['amount']
        avail = [gi for gi in gl_candidates if gi not in matched_gl]

        if not allow_multi:
            if bank_amt >= 0:
                # Bank positive → GL debit
                for gi in avail:
                    g = all_gl[gi]
                    if g['debit'] > 0 and amt_eq(g['debit'], bank_amt):
                        matched_bank[bank_idx] = {'gl_indices': [gi], 'rule': rule}
                        matched_gl.add(gi)
                        return True
            else:
                # Bank negative → GL credit
                for gi in avail:
                    g = all_gl[gi]
                    if g['credit'] > 0 and amt_eq(g['credit'], abs(bank_amt)):
                        matched_bank[bank_idx] = {'gl_indices': [gi], 'rule': rule}
                        matched_gl.add(gi)
                        return True
        return False

    # ---- RULE 1: Check number ----
    for bi in regular_bank:
        b = all_bank[bi]
        if b['check_number']:
            chk = b['check_number']
            cands = [gi for gi, g in enumerate(all_gl)
                     if g['reference'] and g['reference'].strip() == chk]
            try_match(bi, cands, f'Check #{chk}')

    # ---- RULE 2: PAYLEASE.COM → "PAYLEASE" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'PAYLEASE.COM' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'PAYLEASE' in g['remarks'].upper()]
            try_match(bi, cands, 'PAYLEASE')

    # ---- RULE 3: Bay City MHC LLC Settlement → "CKS" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        du = b['description'].upper()
        if 'BAY CITY MHC LLC' in du and 'SETTLEMENT' in du:
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'CKS' in g['remarks'].upper()]
            try_match(bi, cands, 'CKS (Bay City MHC LLC Settlement)')

    # ---- RULE 4: Bay City Settlement + Lakeshore Management → ACH/EFT Deposit ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        du = b['description'].upper()
        if 'BAY CITY SETTLEMENT' in du and 'LAKESHORE MANAGEMENT' in du:
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'ACH DEPOSIT' in g['remarks'].upper()
                        or 'EFT DEPOSIT' in g['remarks'].upper()]
            try_match(bi, cands, 'ACH/EFT Deposit (Bay City Settlement)')

    # ---- RULE 5: BNKCD → "CC Deposit" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'BNKCD' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'CC DEPOSIT' in g['remarks'].upper()]
            try_match(bi, cands, 'CC Deposit (BNKCD)')

    # ---- RULE 6: INTELLIPAY BILLING → "Convenient Payments" ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'INTELLIPAY BILLING' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'CONVENIENT PAYMENTS' in g['remarks'].upper()]
            try_match(bi, cands, 'Convenient Payments (INTELLIPAY)')

    # ---- RULE 7: LAKESHORE EMPLOYMENT → group GL by ref where person_desc has "LSE" ----
    #   Multiple GL lines per payroll period sum to one bank wire
    payroll_bank = [bi for bi in regular_bank
                    if bi not in matched_bank
                    and 'LAKESHORE EMPLOYMENT' in all_bank[bi]['description'].upper()]

    lse_gl = [gi for gi, g in enumerate(all_gl)
              if 'LSE' in g['person_desc'].upper() and gi not in matched_gl]

    # Group by reference (payroll batch)
    lse_groups = defaultdict(list)
    for gi in lse_gl:
        g = all_gl[gi]
        key = g['reference'] if g['reference'] else g['control']
        lse_groups[key].append(gi)

    # Net credit per group
    for ref_key, group_gi in lse_groups.items():
        avail_gi = [gi for gi in group_gi if gi not in matched_gl]
        if not avail_gi:
            continue
        net_credit = sum(all_gl[gi]['credit'] - all_gl[gi]['debit'] for gi in avail_gi)
        for bi in payroll_bank:
            if bi in matched_bank: continue
            if amt_eq(abs(all_bank[bi]['amount']), net_credit):
                matched_bank[bi] = {'gl_indices': avail_gi,
                                     'rule': f'LSE Payroll (LAKESHORE EMPLOYMENT) ref={ref_key}'}
                for gi in avail_gi:
                    matched_gl.add(gi)
                break

    # Also try matching individual LSE GL rows (benefits wires etc.)
    lse_individual = [bi for bi in regular_bank
                      if bi not in matched_bank
                      and 'LAKESHORE EMPLOYMENT' in all_bank[bi]['description'].upper()]
    for bi in lse_individual:
        cands = [gi for gi, g in enumerate(all_gl)
                 if 'LSE' in g['person_desc'].upper() and gi not in matched_gl]
        try_match(bi, cands, 'LSE (LAKESHORE EMPLOYMENT) individual')

    # ---- RULE 8: BILL PAY → "electric" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'BILL PAY' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'ELECTRIC' in g['remarks'].upper()]
            try_match(bi, cands, 'Electric (BILL PAY)')

    # ---- RULE 9: Return → "NSF RP" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'RETURN' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if 'NSF RP' in g['remarks'].upper()
                        or 'NSF RP : ACH DEPOSIT' in g['remarks'].upper()]
            try_match(bi, cands, 'NSF Return')

    # ---- FALLBACK: Amount-only matching for unmatched ----
    remaining_bank = [bi for bi in regular_bank if bi not in matched_bank]
    remaining_gl   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

    bank_by_amt = defaultdict(list)
    for bi in remaining_bank:
        bank_by_amt[round(all_bank[bi]['amount'], 2)].append(bi)

    gl_by_amt = defaultdict(list)
    for gi in remaining_gl:
        gl_by_amt[round(all_gl[gi]['amount'], 2)].append(gi)

    # Unique-amount matches only
    for amt, b_list in sorted(bank_by_amt.items()):
        if len(b_list) == 1 and amt in gl_by_amt and len(gl_by_amt[amt]) == 1:
            bi = b_list[0]
            gi = gl_by_amt[amt][0]
            if bi not in matched_bank and gi not in matched_gl:
                matched_bank[bi] = {'gl_indices': [gi], 'rule': 'Amount-only fallback'}
                matched_gl.add(gi)

    # ============ EXTRA RULES (originally ran after a pickle reload) ============

    # ---- EXTRA RULE A: WT LAKESHORE MANAGEMENT → WC/Audit in GL remarks ----
    # Bank wires to Lakeshore Management for insurance/audit → prev GL WC entries
    prev_label = prev_gl[0]['source'] if prev_gl else None
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        du = b['description'].upper()
        if 'LAKESHORE MANAGEMENT' in du:
            # prefer prev-month entries first (were flagged to clear this month)
            for src_filter in [prev_label, current_label]:
                if src_filter is None:
                    continue
                cands = [gi for gi, g in enumerate(all_gl)
                         if gi not in matched_gl
                         and ('WC' in g['remarks'].upper() or 'AUDIT' in g['remarks'].upper())
                         and g['source'] == src_filter]
                if try_match(bi, cands, 'WC/Audit Insurance (WT Lakeshore Mgmt)'):
                    break

    # ---- EXTRA RULE B: DESKTOP CHECK DEPOSIT → GL with "Deposit" in remarks ----
    for bi in regular_bank:
        if bi in matched_bank: continue
        b = all_bank[bi]
        if 'DEPOSIT' in b['description'].upper():
            cands = [gi for gi, g in enumerate(all_gl)
                     if gi not in matched_gl
                     and g['debit'] > 0
                     and 'DEPOSIT' in g['remarks'].upper()
                     and 'ACCRUE' not in g['remarks'].upper()]
            try_match(bi, cands, 'Check Deposit')

    # ---- EXTRA RULE C: Amount-only pass #2 ----
    remaining_bank2 = [bi for bi in regular_bank if bi not in matched_bank]
    remaining_gl2   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

    bank_by_amt2 = {}
    for bi in remaining_bank2:
        amt = round(all_bank[bi]['amount'], 2)
        bank_by_amt2.setdefault(amt, []).append(bi)

    gl_by_amt2 = {}
    for gi in remaining_gl2:
        amt = round(all_gl[gi]['amount'], 2)
        gl_by_amt2.setdefault(amt, []).append(gi)

    for amt, b_list in sorted(bank_by_amt2.items()):
        if len(b_list) == 1 and amt in gl_by_amt2 and len(gl_by_amt2[amt]) == 1:
            bi = b_list[0]
            gi = gl_by_amt2[amt][0]
            if bi not in matched_bank and gi not in matched_gl:
                matched_bank[bi] = {'gl_indices': [gi], 'rule': 'Amount-only fallback (pass 2)'}
                matched_gl.add(gi)

    # ======================== FINAL STATS ========================
    unmatched_bank_idx = [bi for bi in range(len(all_bank))
                          if bi not in matched_bank
                          and 'STAGECOACH SWEEP' not in all_bank[bi]['description'].upper()]
    unmatched_gl_idx   = [gi for gi in range(len(all_gl)) if gi not in matched_gl]

    total_bank_regular = sum(1 for b in all_bank if 'STAGECOACH SWEEP' not in b['description'].upper())
    total_matched_bank = total_bank_regular - len(unmatched_bank_idx)
    total_gl_all       = len(all_gl)
    total_matched_gl   = total_gl_all - len(unmatched_gl_idx)

    pct_bank = 100.0 * total_matched_bank / total_bank_regular if total_bank_regular else 0
    pct_gl   = 100.0 * total_matched_gl   / total_gl_all       if total_gl_all else 0

    print("=" * 60)
    print("FINAL RECONCILIATION SUMMARY (after extra rules)")
    print("=" * 60)
    print(f"Bank rows (excl stagecoach) : {total_bank_regular}")
    print(f"  Matched                   : {total_matched_bank}  ({pct_bank:.1f}%)")
    print(f"  Unmatched                 : {len(unmatched_bank_idx)}")
    print(f"GL rows total               : {total_gl_all}")
    print(f"  Matched                   : {total_matched_gl}  ({pct_gl:.1f}%)")
    print(f"  Unmatched                 : {len(unmatched_gl_idx)}")
    print(f"Stagecoach sweep (separate) : {len(stagecoach_entries)}")

    from collections import Counter
    rule_counts = Counter(info['rule'] for info in matched_bank.values())
    for rule, cnt in rule_counts.most_common():
        print(f"  {rule:50s} : {cnt} bank rows")

    print("\n=== UNMATCHED BANK TRANSACTIONS ===")
    for bi in unmatched_bank_idx:
        b = all_bank[bi]
        print(f"  [{b['source']}] Row {b['row']} | {b['date']} | {b['amount']:>12,.2f} | {b['description'][:80]}")

    print("\n=== UNMATCHED GL TRANSACTIONS ===")
    for gi in unmatched_gl_idx:
        g = all_gl[gi]
        print(f"  [{g['source']}] Row {g['row']} | {g['date']} | D:{g['debit']:>10,.2f} C:{g['credit']:>10,.2f} | rem={g['remarks'][:50]} | pdesc={g['person_desc'][:40]}")

    results = {
        'all_bank': all_bank,
        'all_gl': all_gl,
        'matched_bank': matched_bank,
        'matched_gl': matched_gl,
        'unmatched_bank_idx': unmatched_bank_idx,
        'unmatched_gl_idx': unmatched_gl_idx,
        'stagecoach_entries': stagecoach_entries,
        'n_prev_bank': len(prev_bank),
        'n_prev_gl': len(prev_gl),
        'stats': {
            'total_bank_regular': total_bank_regular,
            'total_matched_bank': total_matched_bank,
            'total_gl_all': total_gl_all,
            'total_matched_gl': total_matched_gl,
            'pct_bank': pct_bank,
            'pct_gl': pct_gl,
            'total_stagecoach': len(stagecoach_entries),
        }
    }

    if pkl_out:
        with open(pkl_out, 'wb') as f:
            pickle.dump(results, f)
        print(f"Final data saved → {pkl_out}")

    return results
