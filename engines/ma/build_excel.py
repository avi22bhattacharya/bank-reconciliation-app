"""
Build the Manage America bank reconciliation output workbook (refactored
from "Manage America Bank recon/build_excel.py"; sheet construction
unchanged).

Sheets: Summary, Bank Statement, GL (trailing space), Un-Reconcile transactions

Codes: 1/-1 unreconciled, 2/-2 same-month match, 3/-3 cross-month match,
4/-4 Stagecoach Sweep pair legs.

Changes vs the original: the CONFIG block (paths, property strings, period,
beginning balance, PREV_GL_COUNT) became run() parameters; "bank index 0 is
the prior-month entry" generalized to n_prev_bank.
"""
import calendar
from collections import Counter
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font

CUR = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
DATEFMT = 'mm/dd/yyyy'


def parse_date(d):
    s = str(d)[:10]
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def amt_eq(a, b, tol=0.005):
    return abs(float(a) - float(b)) < tol


def assign_codes(data: dict, n_prev_gl: int | None = None,
                 n_prev_bank: int | None = None):
    """Sweep pairing + reconciliation codes for every bank/GL record.

    Returns (bank_code, gl_code, gl_reconcile_date) where codes are:
    1/-1 unreconciled, 2/-2 same-month match, 3/-3 cross-month match,
    4/-4 Stagecoach Sweep pair legs (unpaired sweeps stay 1/-1 and carry
    forward as outstanding items). Shared by the workbook writer and the
    DB persistence layer so both see the same paired/unpaired split.
    """
    all_bank = data['all_bank']
    all_gl = data['all_gl']
    matched_bank = data['matched_bank']

    if n_prev_gl is None:
        n_prev_gl = data.get('n_prev_gl', 0)
    if n_prev_bank is None:
        n_prev_bank = data.get('n_prev_bank', 0)

    # ============================================================
    # STEP 1: Stagecoach Sweep bank-to-bank pairing (FIFO, exact amount)
    # ============================================================
    sweep_debits, sweep_credits = [], []
    for bi, b in enumerate(all_bank):
        du = b['description'].upper()
        if 'STAGECOACH SWEEP DEBIT' in du:
            sweep_debits.append(bi)
        elif 'STAGECOACH SWEEP CREDIT' in du:
            sweep_credits.append(bi)

    sweep_debits.sort(key=lambda bi: parse_date(all_bank[bi]['date']))
    sweep_credits.sort(key=lambda bi: parse_date(all_bank[bi]['date']))

    sweep_pairs = []
    used_credits = set()
    unpaired_sweep = []
    for d_bi in sweep_debits:
        d_amt = abs(all_bank[d_bi]['amount'])
        d_date = parse_date(all_bank[d_bi]['date'])
        match = None
        for c_bi in sweep_credits:
            if c_bi in used_credits:
                continue
            c_date = parse_date(all_bank[c_bi]['date'])
            if c_date is not None and c_date >= d_date and amt_eq(abs(all_bank[c_bi]['amount']), d_amt):
                match = c_bi
                break
        if match is not None:
            sweep_pairs.append((d_bi, match))
            used_credits.add(match)
        else:
            unpaired_sweep.append(d_bi)

    unpaired_sweep += [c for c in sweep_credits if c not in used_credits]

    print(f"Sweep pairs: {len(sweep_pairs)}, unpaired sweep entries: {len(unpaired_sweep)}")
    for bi in unpaired_sweep:
        print(f"  UNPAIRED bank[{bi}]: {all_bank[bi]['date']} {all_bank[bi]['amount']} {all_bank[bi]['description']}")

    # ============================================================
    # STEP 2: Assign Code (1/-1/2/-2/3/-3/4/-4) and Reconcile Date
    # ============================================================
    bank_code = {}
    gl_code = {}
    gl_reconcile_date = {}

    # 2a: sweep pairs -> -4 (debit leg) / 4 (credit leg)
    for d_bi, c_bi in sweep_pairs:
        bank_code[d_bi] = -4
        bank_code[c_bi] = 4

    # 2b: leftover unpaired sweep -> 1 / -1 by sign
    for bi in unpaired_sweep:
        amt = all_bank[bi]['amount']
        bank_code[bi] = 1 if amt > 0 else -1

    # 2c: regular matched_bank groups (skip the empty-gl_indices stagecoach markers)
    same_month_groups = []
    cross_month_groups = []
    for bi, info in matched_bank.items():
        gl_idxs = info['gl_indices']
        if not gl_idxs:
            continue  # "STAGECOACH SWEEP (internal)" marker - handled in step 1
        b = all_bank[bi]
        is_deposit = b['amount'] > 0
        gl_months = set('prev' if gi < n_prev_gl else 'curr' for gi in gl_idxs)
        bank_month = 'prev' if bi < n_prev_bank else 'curr'
        if gl_months == {bank_month}:
            code = 2 if is_deposit else -2
            same_month_groups.append(bi)
        else:
            # cross-month (bank current / GL prior month, or vice versa)
            code = 3 if is_deposit else -3
            cross_month_groups.append(bi)
        bank_code[bi] = code
        for gi in gl_idxs:
            gl_code[gi] = code
            gl_reconcile_date[gi] = b['date']

    print(f"Same-month matched groups (2/-2): {len(same_month_groups)}")
    print(f"Cross-month matched groups (3/-3): {len(cross_month_groups)}")

    # 2d: unmatched bank rows (not sweep, not matched) -> 1 / -1 by sign
    for bi, b in enumerate(all_bank):
        if bi in bank_code:
            continue
        bank_code[bi] = 1 if b['amount'] > 0 else -1

    # 2e: unmatched GL rows -> 1 / -1 by sign
    for gi, g in enumerate(all_gl):
        if gi in gl_code:
            continue
        gl_code[gi] = 1 if g['debit'] > 0 else -1

    assert len(bank_code) == len(all_bank), (len(bank_code), len(all_bank))
    assert len(gl_code) == len(all_gl), (len(gl_code), len(all_gl))

    print("Bank code distribution:", Counter(bank_code.values()))
    print("GL code distribution:", Counter(gl_code.values()))

    return bank_code, gl_code, gl_reconcile_date


def run(data: dict, output_path: str, *,
        property_name: str, property_title: str, property_code: str,
        gl_account_label: str, wfb_account: str,
        period_start: date, beginning_balance: float,
        bank_ending_balance: float | None = None,
        n_prev_gl: int | None = None, n_prev_bank: int | None = None) -> str:
    """Write the 4-sheet MA output workbook from the reconcile_full results dict."""
    all_bank = data['all_bank']
    all_gl = data['all_gl']

    if n_prev_gl is None:
        n_prev_gl = data.get('n_prev_gl', 0)
    if n_prev_bank is None:
        n_prev_bank = data.get('n_prev_bank', 0)

    last_day = calendar.monthrange(period_start.year, period_start.month)[1]
    period_end = date(period_start.year, period_start.month, last_day)
    date_range_label = (f"Detail Date Range: {period_start.strftime('%m/%d/%y')} - "
                        f"{period_end.strftime('%m/%d/%y')}  (Accural basis)")
    period_label = period_start.strftime("%b '") + period_start.strftime("%y")

    bank_code, gl_code, gl_reconcile_date = assign_codes(data, n_prev_gl, n_prev_bank)

    # ============================================================
    # Build workbook
    # ============================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------
    # Sheet: Bank Statement
    # ------------------------------------------------------------
    ws = wb.create_sheet('Bank Statement')
    ws.append(['Date', 'Amount', 'Check Number', 'Code', 'Description'])
    for c in range(1, 6):
        ws.cell(row=1, column=c).font = Font(bold=True)

    bank_order = sorted(range(len(all_bank)), key=lambda i: (parse_date(all_bank[i]['date']), i))

    r = 2
    for bi in bank_order:
        b = all_bank[bi]
        ws.cell(row=r, column=1, value=parse_date(b['date'])).number_format = DATEFMT
        cell = ws.cell(row=r, column=2, value=b['amount'])
        cell.number_format = CUR
        ws.cell(row=r, column=3, value=b['check_number'] if b['check_number'] else '')
        ws.cell(row=r, column=4, value=bank_code[bi])
        ws.cell(row=r, column=5, value=b['description'])
        r += 1

    last_bank_row = r - 1
    r += 1  # blank spacer
    ws.cell(row=r, column=1, value=1)
    ws.cell(row=r, column=2, value=f'=SUMIF(D2:D{last_bank_row},1,B2:B{last_bank_row})').number_format = CUR
    ws.cell(row=r, column=3, value='Un-Reconciled Deposits')
    r += 2
    ws.cell(row=r, column=1, value=-1)
    ws.cell(row=r, column=2, value=f'=SUMIF(D2:D{last_bank_row},-1,B2:B{last_bank_row})').number_format = CUR
    ws.cell(row=r, column=3, value='Un-Reconciled Withdrawals')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 90

    print(f"Bank Statement: {last_bank_row - 1} data rows, totals at rows {r - 2} and {r}")

    # ------------------------------------------------------------
    # Sheet: GL  (trailing space, matches the reference sheet name)
    # ------------------------------------------------------------
    wsg = wb.create_sheet('GL ')
    wsg['B1'] = '                    General Ledger'
    wsg['B2'] = f'Property: {property_name} ({property_code})'
    wsg['B3'] = date_range_label
    wsg['A4'] = f'{gl_account_label}  (Bank Account)'
    wsg.cell(row=4, column=9, value=beginning_balance).number_format = CUR

    header = ['Property Name', 'Date', 'Control', 'Reference', 'Description', 'Remarks',
              'Debit', 'Credit', 'Balance', 'Reconcile Date', 'Code']
    for c, h in enumerate(header, start=1):
        cell = wsg.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True)

    gl_row_of = {}
    r = 6
    running_balance = beginning_balance
    for gi, g in enumerate(all_gl):
        is_prev = gi < n_prev_gl
        wsg.cell(row=r, column=1, value=property_name)
        dt = parse_date(g['date'])
        wsg.cell(row=r, column=2, value=dt).number_format = DATEFMT
        wsg.cell(row=r, column=3, value=g['control'])
        wsg.cell(row=r, column=4, value=g['reference'])
        wsg.cell(row=r, column=5, value=g['person_desc'])
        wsg.cell(row=r, column=6, value=g['remarks'])
        debit = g['debit'] if g['debit'] else None
        credit = g['credit'] if g['credit'] else None
        if debit is not None:
            wsg.cell(row=r, column=7, value=debit).number_format = CUR
        if credit is not None:
            wsg.cell(row=r, column=8, value=credit).number_format = CUR
        if not is_prev:
            running_balance = running_balance + (g['debit'] or 0) - (g['credit'] or 0)
            wsg.cell(row=r, column=9, value=round(running_balance, 2)).number_format = CUR
        rd = gl_reconcile_date.get(gi)
        if rd:
            wsg.cell(row=r, column=10, value=parse_date(rd)).number_format = DATEFMT
        wsg.cell(row=r, column=11, value=gl_code[gi])
        gl_row_of[gi] = r
        r += 1

    last_gl_row = r - 1
    curr_first_row = gl_row_of[n_prev_gl] if n_prev_gl in gl_row_of else 6
    curr_last_row = gl_row_of[len(all_gl) - 1] if all_gl else 6

    # Totals row (current month only, current-month gl range)
    totals_row = r
    wsg.cell(row=totals_row, column=7, value=f'=SUM(G{curr_first_row}:G{curr_last_row})').number_format = CUR
    wsg.cell(row=totals_row, column=8, value=f'=SUM(H{curr_first_row}:H{curr_last_row})').number_format = CUR
    wsg.cell(row=totals_row, column=9, value=f'=I4+G{totals_row}-H{totals_row}').number_format = CUR

    r = totals_row + 2
    wsg.cell(row=r, column=6, value=1)
    wsg.cell(row=r, column=7, value=f'=+SUMIF(K6:K{last_gl_row},1,G6:G{last_gl_row})+SUMIF(K6:K{last_gl_row},1,H6:H{last_gl_row})').number_format = CUR
    wsg.cell(row=r, column=8, value='Un-Recorded deposits')
    r += 2
    wsg.cell(row=r, column=6, value=-1)
    wsg.cell(row=r, column=7, value=f'=+SUMIF(K6:K{last_gl_row},-1,G6:G{last_gl_row})+SUMIF(K6:K{last_gl_row},-1,H6:H{last_gl_row})').number_format = CUR
    wsg.cell(row=r, column=8, value='Un-Recorded withdrawal')

    for col, w in zip('ABCDEFGHIJK', [11, 11, 12, 16, 45, 45, 13, 13, 13, 13, 6]):
        wsg.column_dimensions[col].width = w

    GL_TOTALS_ROW = totals_row
    print(f"GL: {last_gl_row - 5} data rows (rows 6-{last_gl_row}), totals row {totals_row}")

    # ------------------------------------------------------------
    # Sheet: Un-Reconcile transactions
    # ------------------------------------------------------------
    wsu = wb.create_sheet('Un-Reconcile transactions')

    unmatched_bank = sorted([bi for bi, c in bank_code.items() if c in (1, -1)],
                             key=lambda bi: parse_date(all_bank[bi]['date']))
    unmatched_gl_dep = sorted([gi for gi, c in gl_code.items() if c == 1],
                               key=lambda gi: parse_date(all_gl[gi]['date']))
    unmatched_gl_wd = sorted([gi for gi, c in gl_code.items() if c == -1],
                              key=lambda gi: parse_date(all_gl[gi]['date']))

    r = 1
    wsu.cell(row=r, column=1, value='Bank Statement').font = Font(bold=True)
    r += 1
    bank_hdr = ['Property Name', 'Date', 'Control', 'Reference', 'Descriptions', 'Remarks', 'Amount', 'Comments']
    for c, h in enumerate(bank_hdr, start=1):
        wsu.cell(row=r, column=c, value=h).font = Font(bold=True)
    r += 1
    bank_section_start = r
    for bi in unmatched_bank:
        b = all_bank[bi]
        wsu.cell(row=r, column=1, value=property_name)
        wsu.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
        wsu.cell(row=r, column=3, value=None)
        wsu.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
        wsu.cell(row=r, column=5, value=None)
        wsu.cell(row=r, column=6, value=b['description'])
        cell = wsu.cell(row=r, column=7, value=b['amount'])
        cell.number_format = CUR
        wsu.cell(row=r, column=8, value='AR' if b['amount'] > 0 else 'AP')
        r += 1
    bank_section_end = r - 1
    wsu.cell(row=r, column=7, value=f'=SUM(G{bank_section_start}:G{bank_section_end})').number_format = CUR
    r += 2

    wsu.cell(row=r, column=1, value='General Ledger').font = Font(bold=True)
    r += 1
    gl_hdr = ['Property Name', 'Date', 'Control', 'Reference', 'Description', 'Remarks', 'Amount', 'Comments']
    for c, h in enumerate(gl_hdr, start=1):
        wsu.cell(row=r, column=c, value=h).font = Font(bold=True)
    r += 1

    def write_gl_section(rows_idx, r):
        start = r
        for gi in rows_idx:
            g = all_gl[gi]
            wsu.cell(row=r, column=1, value=property_name)
            wsu.cell(row=r, column=2, value=parse_date(g['date'])).number_format = DATEFMT
            wsu.cell(row=r, column=3, value=g['control'])
            wsu.cell(row=r, column=4, value=g['reference'])
            wsu.cell(row=r, column=5, value=g['person_desc'])
            wsu.cell(row=r, column=6, value=g['remarks'])
            amount = g['debit'] if g['debit'] else -g['credit']
            cell = wsu.cell(row=r, column=7, value=amount)
            cell.number_format = CUR
            wsu.cell(row=r, column=8, value='AR' if amount > 0 else 'AP')
            r += 1
        end = r - 1
        if end >= start:
            wsu.cell(row=r, column=7, value=f'=SUM(G{start}:G{end})').number_format = CUR
            r += 1
        return start, end, r

    dep_start, dep_end, r = write_gl_section(unmatched_gl_dep, r)
    r += 1
    wd_start, wd_end, r = write_gl_section(unmatched_gl_wd, r)

    for col, w in zip('ABCDEFGH', [11, 11, 14, 18, 50, 50, 13, 9]):
        wsu.column_dimensions[col].width = w

    print(f"Un-Reconcile transactions: bank rows {bank_section_start}-{bank_section_end}, "
          f"GL deposits {dep_start}-{dep_end}, GL withdrawals {wd_start}-{wd_end}")

    # ------------------------------------------------------------
    # Sheet: Summary
    # ------------------------------------------------------------
    wss = wb.create_sheet('Summary', 0)  # make it first sheet

    wss['B1'] = ' '
    wss['B2'] = f'Bank Reconciliation - {property_title}'
    wss['B3'] = wfb_account
    wss['B4'] = period_start
    wss['B4'].number_format = DATEFMT
    wss['B2'].font = Font(bold=True)
    wss['B5'] = 'Bank Information'
    wss['B5'].font = Font(bold=True)

    r = 7
    wss.cell(row=r, column=6, value='Balance per bank at the month end')
    wss.cell(row=r, column=8, value=bank_ending_balance if bank_ending_balance is not None else 0).number_format = CUR
    H7_row = r

    r += 2
    wss.cell(row=r, column=2, value='Outstanding Bank Deposits:').font = Font(italic=True)
    r += 1
    items_first_row = r
    ub_dep = sorted([bi for bi, c in bank_code.items() if c == 1], key=lambda bi: parse_date(all_bank[bi]['date']))
    ub_wd = sorted([bi for bi, c in bank_code.items() if c == -1], key=lambda bi: parse_date(all_bank[bi]['date']))

    for bi in ub_dep:
        b = all_bank[bi]
        wss.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
        wss.cell(row=r, column=3, value='AR')
        wss.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
        wss.cell(row=r, column=5, value=b['description'])
        wss.cell(row=r, column=7, value=b['amount']).number_format = CUR
        r += 1

    r += 1
    wss.cell(row=r, column=2, value='Outstanding Bank Disbursement:').font = Font(italic=True)
    r += 1
    for bi in ub_wd:
        b = all_bank[bi]
        wss.cell(row=r, column=2, value=parse_date(b['date'])).number_format = DATEFMT
        wss.cell(row=r, column=3, value='AP')
        wss.cell(row=r, column=4, value=b['check_number'] if b['check_number'] else '')
        wss.cell(row=r, column=5, value=b['description'])
        wss.cell(row=r, column=7, value=b['amount']).number_format = CUR
        r += 1

    items_last_row = r - 1
    r += 1
    wss.cell(row=r, column=6, value='Outstanding Items Total')
    wss.cell(row=r, column=8, value=f'=+SUM(G{items_first_row}:G{items_last_row})').number_format = CUR
    H16_row = r

    r += 2
    wss.cell(row=r, column=6, value='Adjusted bank balance')
    wss.cell(row=r, column=8, value=f'=+H{H7_row}-H{H16_row}').number_format = CUR
    H18_row = r

    r += 2
    wss.cell(row=r, column=2, value=f'General Ledger Information - {gl_account_label}').font = Font(bold=True)

    r += 2
    wss.cell(row=r, column=6, value=' Balance per TB at the month end')
    wss.cell(row=r, column=8, value=f"=+'GL '!I{GL_TOTALS_ROW}").number_format = CUR
    H22_row = r

    r += 2
    wss.cell(row=r, column=2, value='Un-reconcile Outstanding Deposits:').font = Font(italic=True)
    r += 1
    outstanding_dep_first_row = r
    gl_dep_outstanding = sorted([gi for gi, c in gl_code.items() if c == 1 and gi < n_prev_gl],
                                 key=lambda gi: parse_date(all_gl[gi]['date']))
    gl_dep_curr = sorted([gi for gi, c in gl_code.items() if c == 1 and gi >= n_prev_gl],
                          key=lambda gi: parse_date(all_gl[gi]['date']))
    gl_wd_outstanding = sorted([gi for gi, c in gl_code.items() if c == -1 and gi < n_prev_gl],
                                key=lambda gi: parse_date(all_gl[gi]['date']))
    gl_wd_curr = sorted([gi for gi, c in gl_code.items() if c == -1 and gi >= n_prev_gl],
                         key=lambda gi: parse_date(all_gl[gi]['date']))

    def write_summary_gl_items(rows_idx, r):
        for gi in rows_idx:
            g = all_gl[gi]
            amount = g['debit'] if g['debit'] else -g['credit']
            wss.cell(row=r, column=2, value=parse_date(g['date'])).number_format = DATEFMT
            wss.cell(row=r, column=3, value='AR' if amount > 0 else 'AP')
            wss.cell(row=r, column=4, value=g['control'])
            wss.cell(row=r, column=5, value=g['remarks'])
            wss.cell(row=r, column=7, value=amount).number_format = CUR
            r += 1
        return r

    r = write_summary_gl_items(gl_dep_outstanding, r)

    r += 1
    wss.cell(row=r, column=2, value=f"Un-reconcile Deposits: {period_label}").font = Font(italic=True)
    r += 1
    r = write_summary_gl_items(gl_dep_curr, r)
    dep_last_row = r - 1
    r += 1
    wss.cell(row=r, column=6, value='Total ')
    G37_row = r
    wss.cell(row=r, column=7, value=f'=SUM(G{outstanding_dep_first_row}:G{dep_last_row})').number_format = CUR

    r += 2
    wss.cell(row=r, column=2, value='Un-reconciled Outstanding Disbursement:').font = Font(italic=True)
    r += 1
    wd_outstanding_first_row = r
    r = write_summary_gl_items(gl_wd_outstanding, r)

    r += 1
    wss.cell(row=r, column=2, value=f"Un-reconcile Disbursement: {period_label}").font = Font(italic=True)
    r += 1
    r = write_summary_gl_items(gl_wd_curr, r)
    wd_last_row = r - 1
    r += 1
    wss.cell(row=r, column=6, value='Total ')
    G59_row = r
    wss.cell(row=r, column=7, value=f'=SUM(G{wd_outstanding_first_row}:G{wd_last_row})').number_format = CUR

    r += 2
    wss.cell(row=r, column=6, value='Sub-Total ')
    G61_row = r
    wss.cell(row=r, column=7, value=f'=+G{G59_row}+G{G37_row}').number_format = CUR

    r += 2
    wss.cell(row=r, column=6, value='Adjusted book balance')
    wss.cell(row=r, column=8, value=f'=+H{H22_row}-G{G61_row}').number_format = CUR
    H63_row = r

    r += 2
    wss.cell(row=r, column=6, value='Variance (s/b zero)')
    wss.cell(row=r, column=8, value=f'=+H{H18_row}-H{H63_row}').number_format = CUR

    for col, w in zip('BCDEFGH', [12, 6, 16, 55, 32, 14, 14]):
        wss.column_dimensions[col].width = w

    print(f"Summary sheet built, {r} rows total")

    wb.save(output_path)
    print(f"Saved {output_path}")
    return output_path
