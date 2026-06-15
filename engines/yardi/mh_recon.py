"""
MH Reconciliation – Deposit Number Lookup (refactored from
deposit-register-trial/mh_recon.py; matching logic unchanged).

Reads the Deposit Register and one or more GL source sheets, then writes
an Excel file with three new columns appended to each GL sheet:

  • Tenant Code   – extracted from the Person/Description field
  • Unique ID     – Tenant Code + Date + Remarks (the match key)
  • Deposit Number – digits-only deposit number from the Deposit Register

Changes vs the original:
  - run(deposit_register, gl_sources, output_path) replaces module CONFIG;
    every source dict carries an absolute "file" path.
  - LibreOffice .xls conversion removed: pandas reads .xls via xlrd directly.
  - Deposit Register header row auto-detected (row containing "Tenant Code")
    when "header_row" is omitted from the config.
  - safe() also handles pd.NA / NaT (pandas 3.0).
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def safe(v):
    """Return a value safe for openpyxl (no NaN/NA, no Series)."""
    if v is None:
        return ""
    if isinstance(v, pd.Series):
        return str(v.iloc[0]) if len(v) > 0 else ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return v


def extract_tenant_code(text):
    """'Smith (t0000771)' → 't0000771', else ''."""
    m = re.search(r'\(([tTrR]\d+)\)', str(text))
    return m.group(1).lower() if m else ""


def digits_only(raw):
    """'ACH-RC 37075' → '37075';  'CC-874' → '874';  '1000' → '1000'."""
    d = re.sub(r'[^0-9]', '', str(raw))
    return d if d else str(raw)


def detect_header_row(raw, marker="Tenant Code", max_scan=30):
    """Find the 0-based row index whose cells contain the marker column name."""
    for i in range(min(max_scan, len(raw))):
        if any(str(v).strip() == marker for v in raw.iloc[i].tolist()):
            return i
    raise ValueError(f"Could not find a header row containing {marker!r}")


def build_lookup(cfg):
    """Build {(tenant_code, date, notes): raw_deposit_number} from Deposit Register."""
    raw = pd.read_excel(cfg["file"], sheet_name=cfg["sheet"], header=None)
    hr = cfg.get("header_row")
    if hr is None:
        hr = detect_header_row(raw)
    raw.columns = raw.iloc[hr].tolist()
    df = raw.iloc[hr + 1:]
    df = df[df["Tenant Code"].notna()].copy()
    df["_tc"]    = df["Tenant Code"].astype(str).str.strip().str.lower()
    df["_date"]  = pd.to_datetime(df["Received Date"], errors="coerce").dt.date
    df["_notes"] = df["Notes"].astype(str).str.strip().str.lower()
    lookup = {}
    for _, row in df.iterrows():
        key = (row["_tc"], row["_date"], row["_notes"])
        if key not in lookup and pd.notna(row["Deposit Number"]):
            lookup[key] = str(row["Deposit Number"]).strip()
    print(f"  Deposit Register loaded: {len(df)} entries, {len(lookup)} unique lookup keys")
    return lookup


def do_lookup(lookup, person, date_val, remarks):
    """Return (tenant_code, unique_id, deposit_number_clean) for one GL row."""
    tc = extract_tenant_code(person)
    if not tc:
        return "", "", ""
    try:
        date = pd.to_datetime(date_val).date()
    except Exception:
        return tc, "", ""
    rem = str(remarks).strip().lower()
    uid = tc + str(date) + rem
    raw = lookup.get((tc, date, rem))
    dep = digits_only(raw) if raw else ""
    return tc, uid, dep


def process_source(cfg, lookup):
    raw = pd.read_excel(cfg["file"], sheet_name=cfg["sheet"], header=None)
    hr  = cfg["header_row"]
    col_names = [str(v) if pd.notna(v) else "" for v in raw.iloc[hr].tolist()]

    # Deduplicate column names (e.g. two "Description" columns)
    seen, deduped = {}, []
    for c in col_names:
        if c in seen:
            seen[c] += 1
            deduped.append(f"{c} ({seen[c]})")
        else:
            seen[c] = 0
            deduped.append(c)
    col_names = deduped

    meta_rows = [raw.iloc[i].tolist() for i in range(hr)]   # rows above header
    data_rows = []

    p_idx = cfg["person_col_idx"]
    d_idx = cfg["date_col_idx"]
    r_idx = cfg["remarks_col_idx"]

    for i in range(hr + 1, len(raw)):
        row = raw.iloc[i]
        person  = safe(row.iloc[p_idx])
        date_v  = safe(row.iloc[d_idx])
        remarks = safe(row.iloc[r_idx])
        tc, uid, dep = do_lookup(lookup, person, date_v, remarks)
        vals = [safe(row.iloc[c]) for c in range(len(row))] + [tc, uid, dep]
        data_rows.append(vals)

    matched = sum(1 for r in data_rows if r[-1])
    print(f"  [{cfg['label']}] {matched}/{len(data_rows)} rows matched to a Deposit Number")
    return meta_rows, col_names, data_rows


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

BLUE  = PatternFill("solid", fgColor="1F4E79")
BFNT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
YEL   = PatternFill("solid", fgColor="FFD966")
YFNT  = Font(bold=True, color="000000", name="Calibri", size=10)
LYEL  = PatternFill("solid", fgColor="FFFACD")
DFNT  = Font(name="Calibri", size=10)
CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
NEW_COLS = ["Tenant Code", "Unique ID", "Deposit Number"]


def write_gl_sheet(wb, name, meta_rows, col_names, data_rows):
    ws = wb.create_sheet(name)

    for meta in meta_rows:
        ws.append([safe(v) for v in meta])

    full_header = col_names + NEW_COLS
    ws.append(full_header)
    hr = ws.max_row
    for c, h in enumerate(full_header, 1):
        cell = ws.cell(row=hr, column=c)
        if h in NEW_COLS:
            cell.fill = YEL;  cell.font = YFNT
        else:
            cell.fill = BLUE; cell.font = BFNT
        cell.alignment = CTR

    n = len(full_header)
    for row_vals in data_rows:
        ws.append(row_vals)
        dr = ws.max_row
        for c in range(1, n + 1):
            ws.cell(row=dr, column=c).font = DFNT
        ws.cell(row=dr, column=n).fill = LYEL   # highlight Deposit Number

    # Column widths
    for c in range(1, n + 1):
        col_l = get_column_letter(c)
        w = len(str(full_header[c - 1])) + 2
        for r in range(hr, min(hr + 80, ws.max_row + 1)):
            v = ws.cell(row=r, column=c).value
            if v:
                w = max(w, len(str(v)) + 2)
        ws.column_dimensions[col_l].width = min(w, 45)


# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────

def run(deposit_register: dict, gl_sources: list[dict], output_path: str) -> list[tuple]:
    """Enrich GL sheets with deposit numbers.

    deposit_register: {"file": <abs path>, "sheet": str, "header_row": int|None}
    gl_sources: list of {"label", "file" (abs path), "sheet", "header_row",
                         "person_col_idx", "date_col_idx", "remarks_col_idx"}
    Returns [(label, total_rows, matched_rows), ...].
    """
    print("Loading Deposit Register...")
    lookup = build_lookup(deposit_register)

    wb = Workbook()
    wb.remove(wb.active)

    all_stats = []
    for cfg in gl_sources:
        print(f"\nProcessing: {cfg['file']} → sheet '{cfg['sheet']}'")
        meta, cols, data = process_source(cfg, lookup)
        write_gl_sheet(wb, cfg["label"], meta, cols, data)
        matched = sum(1 for r in data if r[-1])
        all_stats.append((cfg["label"], len(data), matched))

    # Summary sheet
    ws_s = wb.create_sheet("Summary")
    ws_s.append(["Sheet", "Total Rows", "Deposit # Matched", "Unmatched"])
    for label, total, matched in all_stats:
        ws_s.append([label, total, matched, total - matched])

    wb.save(output_path)
    print(f"\nOutput saved: {output_path}")
    return all_stats
