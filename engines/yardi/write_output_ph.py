"""
Generate Yardi Bank Reconciliation Output Excel (refactored from
deposit-register-trial/write_output_ph.py; sheet construction unchanged).

4 sheets matching the reference format:
  1. Summary
  2. Bank Statement
  3. GL
  4. Un-Reconcile transactions

Changes vs the original: run(results, output_path, prop) replaces the
module-level JSON path and the 4 hardcoded PH Davie constants; `prop` is a
core.db.PropertyMeta (property_name, account_info, gl_account_id,
prop_label_bank, prop_display).
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from collections import OrderedDict
import calendar

# ── Styles ────────────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_date2(s):
    """Tolerant date parser."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    s = str(s).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

def fmt_date(d):
    if isinstance(d, datetime):
        return d.strftime("%m/%d/%Y")
    if d:
        return str(d)
    return ""

def period_label(d):
    """Returns e.g. "Apr '26" from a datetime or date-string."""
    d = parse_date2(d)
    if not d:
        return ""
    return d.strftime("%b '") + d.strftime("%y")

_INTERNAL_RULES = {"INTERNAL – Stagecoach Sweep", "Contra - Bank"}

# ── Code helpers ──────────────────────────────────────────────────────────────
def bank_code(b, gl_by_id):
    rule = b.get("match_rule", "")
    amt  = b.get("amount", 0) or 0
    sign = 1 if amt >= 0 else -1
    if rule in _INTERNAL_RULES:
        return sign * 4
    if rule:  # matched to GL
        bank_d = parse_date2(b.get("date"))
        for gid in (b.get("match_ids") or []):
            g = gl_by_id.get(gid)
            if g:
                gl_d = parse_date2(g.get("date"))
                if gl_d and bank_d and (gl_d.year, gl_d.month) != (bank_d.year, bank_d.month):
                    return sign * 3
        return sign * 2
    return sign * 1  # unmatched

def gl_code(g, bank_by_id):
    rule = g.get("match_rule", "")
    is_debit = (g.get("debit") or 0) > 0
    sign = 1 if is_debit else -1
    if rule == "Contra - GL":
        return sign * 4
    if g.get("matched"):
        gl_d = parse_date2(g.get("date"))
        for mid in (g.get("match_ids") or []):
            b = bank_by_id.get(mid)
            if b:
                bank_d = parse_date2(b.get("date"))
                if gl_d and bank_d and (gl_d.year, gl_d.month) != (bank_d.year, bank_d.month):
                    return sign * 3
        return sign * 2
    return sign * 1

def gl_reconcile_date(g, bank_by_id):
    rule = g.get("match_rule", "")
    if rule == "Contra - GL":
        return fmt_date(parse_date2(g.get("date")))
    if g.get("matched"):
        for mid in (g.get("match_ids") or []):
            b = bank_by_id.get(mid)
            if b:
                return b.get("date") or ""
    return None

# ── Unreconciled item grouping ────────────────────────────────────────────────
def group_unrec_gl(unrec_items):
    """
    Groups unreconciled GL items for the Summary sheet.
    - K-entries: group by Reference (check number), one row per ref
    - J/R/other entries: group by (Control, Property), one row per combo
    Returns list of dicts: {date, type, d_key, desc, amount}
    """
    groups = OrderedDict()
    for g in sorted(unrec_items, key=lambda x: (parse_date2(x.get("date")) or datetime.min, x.get("control", "") or "")):
        ctl   = g.get("control", "") or ""
        ref   = g.get("ref",     "") or ""
        prop  = g.get("property","") or ""
        is_debit = (g.get("debit") or 0) > 0

        if ctl.upper().startswith("K-"):
            gkey = ("K", ref)
        else:
            gkey = ("JR", ctl, prop)

        if gkey not in groups:
            groups[gkey] = {"items": [], "is_debit": is_debit, "ctl": ctl, "ref": ref}
        groups[gkey]["items"].append(g)

    result = []
    for gkey, gdata in groups.items():
        items = sorted(gdata["items"], key=lambda x: parse_date2(x.get("date")) or datetime.min)
        first = items[0]
        is_debit = gdata["is_debit"]

        total = round(
            sum((g.get("debit") or 0) - (g.get("credit") or 0) for g in items), 2
        )

        if gkey[0] == "K":
            # AP check: use Description (vendor name) for E column
            e_desc = first.get("desc", "") or first.get("remarks", "") or ""
            d_key  = gdata["ref"]    # Reference = check number
        else:
            # J/R entries: use Remarks for E column
            e_desc = first.get("remarks", "") or first.get("desc", "") or ""
            d_key = gdata["ctl"]     # Control (J-xxxxx or R-xxxxx)

        result.append({
            "date":   parse_date2(first.get("date")),
            "type":   "AR" if is_debit else "AP",
            "d_key":  d_key,
            "desc":   e_desc,
            "amount": total,         # positive for deposits, negative for withdrawals
        })
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_summary(wb, r, prop):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 2, "B": 15, "C": 6, "D": 18, "E": 40,
                  "F": 26, "G": 18, "H": 18, "I": 52}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    def wc(row, col, val, bold=False, italic=False, size=11,
           halign=None, color="000000", num_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, italic=italic, size=size, color=color)
        if halign:
            c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt
        return c

    # ── Header ────────────────────────────────────────────────────────────────
    wc(2, 2, f"Bank Reconciliation - {prop.property_name}", bold=True, size=14)
    wc(3, 2, prop.account_info)

    # Period date = first day of the reporting month (derived from all_bank dates)
    bank_dates = [parse_date2(b.get("date")) for b in r["all_bank"] if b.get("date")]
    bank_dates = [d for d in bank_dates if d]
    if bank_dates:
        max_d = max(bank_dates)
        period_date = datetime(max_d.year, max_d.month, 1)
    else:
        period_date = None
    ws.cell(row=4, column=2, value=period_date)
    ws.cell(row=4, column=2).number_format = "MM/DD/YYYY"

    wc(5, 2, "Bank Information", bold=True)

    # ── Bank section ─────────────────────────────────────────────────────────
    wc(7, 6, "Balance per bank at the month end")

    sc_ids    = set(r.get("stagecoach_ids", []))
    all_bank  = r["all_bank"]
    all_gl    = r["all_gl"]

    # Unreconciled bank entries (not internal/stagecoach)
    unrec_bank = [b for b in all_bank
                  if not b.get("match_rule") and b["id"] not in sc_ids]
    unrec_bank_deposits    = [b for b in unrec_bank if (b.get("amount") or 0) > 0]
    unrec_bank_withdrawals = [b for b in unrec_bank if (b.get("amount") or 0) < 0]

    total_unrec_bank_dep  = round(sum(b.get("amount",0) for b in unrec_bank_deposits),  2)
    total_unrec_bank_with = round(sum(b.get("amount",0) for b in unrec_bank_withdrawals), 2)
    total_outstanding_bank = round(total_unrec_bank_dep + total_unrec_bank_with, 2)

    # GL ending balance = opening + sum(debits) - sum(credits)
    gl_opening = r.get("gl_opening_balance", 0.0) or 0.0
    total_gl_dr = round(sum(g.get("debit",  0) or 0 for g in all_gl), 2)
    total_gl_cr = round(sum(g.get("credit", 0) or 0 for g in all_gl), 2)
    gl_ending   = round(gl_opening + total_gl_dr - total_gl_cr, 2)

    # Unreconciled GL items
    unrec_gl = [g for g in all_gl if not g.get("matched")]
    prev_unrec = [g for g in unrec_gl if "Unrec" in (g.get("source") or "")]
    curr_unrec = [g for g in unrec_gl if "Unrec" not in (g.get("source") or "")]

    prev_unrec_dep = [g for g in prev_unrec if (g.get("debit") or 0) > 0]
    prev_unrec_wit = [g for g in prev_unrec if (g.get("credit") or 0) > 0]
    curr_unrec_dep = [g for g in curr_unrec if (g.get("debit") or 0) > 0]
    curr_unrec_wit = [g for g in curr_unrec if (g.get("credit") or 0) > 0]

    prev_dep_groups = group_unrec_gl(prev_unrec_dep)
    curr_dep_groups = group_unrec_gl(curr_unrec_dep)
    prev_wit_groups = group_unrec_gl(prev_unrec_wit)
    curr_wit_groups = group_unrec_gl(curr_unrec_wit)

    total_unrec_gl_dep = round(sum(g["amount"] for g in prev_dep_groups + curr_dep_groups), 2)
    total_unrec_gl_wit = round(sum(g["amount"] for g in prev_wit_groups + curr_wit_groups), 2)
    gl_sub_total       = round(total_unrec_gl_dep + total_unrec_gl_wit, 2)

    adjusted_book  = round(gl_ending - gl_sub_total, 2)
    bank_ending    = round(adjusted_book + total_outstanding_bank, 2)
    adjusted_bank  = round(bank_ending - total_outstanding_bank, 2)
    variance       = round(adjusted_bank - adjusted_book, 2)

    # H7: bank balance
    c = ws.cell(row=7, column=8, value=bank_ending)
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    c.alignment = Alignment(horizontal="right")

    # Outstanding bank deposits section
    wc(9, 2, "Outstanding Bank Deposits:", bold=True)

    dep_data_start = 11
    dep_data_end   = dep_data_start - 1
    for b in sorted(unrec_bank_deposits, key=lambda x: parse_date2(x.get("date")) or datetime.min):
        row = dep_data_end + 1
        dep_data_end = row
        ws.cell(row=row, column=2, value=b.get("date") or "")
        ws.cell(row=row, column=3, value="AR")
        ws.cell(row=row, column=5, value=b.get("description", ""))
        c = ws.cell(row=row, column=7, value=b.get("amount", 0))
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    total_dep_row = dep_data_end + 2
    wc(total_dep_row, 6, "Total")
    c = ws.cell(row=total_dep_row, column=7, value=total_unrec_bank_dep)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    # Outstanding bank withdrawals section
    dis_hdr_row = total_dep_row + 1
    wc(dis_hdr_row, 2, "Outstanding Bank Disbursement:", bold=True)
    dis_data_start = dis_hdr_row + 2
    dis_data_end   = dis_data_start - 1

    for b in sorted(unrec_bank_withdrawals, key=lambda x: parse_date2(x.get("date")) or datetime.min):
        row = dis_data_end + 1
        dis_data_end = row
        ws.cell(row=row, column=2, value=b.get("date") or "")
        ws.cell(row=row, column=3, value="AP")
        ws.cell(row=row, column=5, value=b.get("description", ""))
        c = ws.cell(row=row, column=7, value=b.get("amount", 0))
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    total_wit_row = dis_data_end + 2
    wc(total_wit_row, 6, "Total")
    c = ws.cell(row=total_wit_row, column=7, value=total_unrec_bank_with)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    outstanding_row = total_wit_row + 2
    wc(outstanding_row, 6, "Outstanding items total")
    c = ws.cell(row=outstanding_row, column=8, value=total_outstanding_bank)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    adj_bank_row = outstanding_row + 2
    wc(adj_bank_row, 6, "Adjusted bank balance", bold=True)
    c = ws.cell(row=adj_bank_row, column=8, value=adjusted_bank)
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    c.font = Font(bold=True)

    # ── GL section ────────────────────────────────────────────────────────────
    gl_hdr_row = adj_bank_row + 2
    gl_acct_prefix = prop.gl_account_id.split()[0] if prop.gl_account_id else ""
    wc(gl_hdr_row, 2, f"General Ledger Information - {gl_acct_prefix} {prop.property_name}", bold=True)

    gl_bal_row = gl_hdr_row + 2
    wc(gl_bal_row, 6, " Balance per TB at the month end")
    c = ws.cell(row=gl_bal_row, column=8, value=gl_ending)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    p_label = period_label(period_date) if period_date else ""

    # ── Prior-period outstanding deposits ────────────────────────────────────
    prev_dep_hdr = gl_bal_row + 2
    wc(prev_dep_hdr, 2, "Un-reconcile Outstanding Deposits:", bold=True)

    prev_dep_data_end = prev_dep_hdr + 1
    for item in prev_dep_groups:
        row = prev_dep_data_end + 1
        prev_dep_data_end = row
        ws.cell(row=row, column=2, value=item["date"])
        if isinstance(item["date"], datetime):
            ws.cell(row=row, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=row, column=3, value=item["type"])
        ws.cell(row=row, column=4, value=item["d_key"])
        ws.cell(row=row, column=5, value=item["desc"])
        c = ws.cell(row=row, column=7, value=item["amount"])
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    # ── Current deposits ──────────────────────────────────────────────────────
    curr_dep_hdr = prev_dep_data_end + 2
    wc(curr_dep_hdr, 2, "Un-reconcile Deposits:", bold=True)
    ws.cell(row=curr_dep_hdr, column=4, value=p_label)
    curr_dep_data_start = curr_dep_hdr + 2
    curr_dep_data_end   = curr_dep_data_start - 1

    for item in curr_dep_groups:
        row = curr_dep_data_end + 1
        curr_dep_data_end = row
        ws.cell(row=row, column=2, value=item["date"])
        if isinstance(item["date"], datetime):
            ws.cell(row=row, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=row, column=3, value=item["type"])
        ws.cell(row=row, column=4, value=item["d_key"])
        ws.cell(row=row, column=5, value=item["desc"])
        c = ws.cell(row=row, column=7, value=item["amount"])
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    total_gl_dep_row = curr_dep_data_end + 2
    wc(total_gl_dep_row, 6, "Total")
    c = ws.cell(row=total_gl_dep_row, column=7, value=total_unrec_gl_dep)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    # ── Prior-period outstanding disbursements ────────────────────────────────
    ws.cell(row=total_gl_dep_row + 1, column=1, value=" ")
    prev_wit_hdr = total_gl_dep_row + 1
    wc(prev_wit_hdr, 2, "Un-reconciled Outstanding Disbursement:", bold=True)

    prev_wit_data_end = prev_wit_hdr + 1
    for item in prev_wit_groups:
        row = prev_wit_data_end + 1
        prev_wit_data_end = row
        ws.cell(row=row, column=2, value=item["date"])
        if isinstance(item["date"], datetime):
            ws.cell(row=row, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=row, column=3, value=item["type"])
        ws.cell(row=row, column=4, value=item["d_key"])
        ws.cell(row=row, column=5, value=item["desc"])
        c = ws.cell(row=row, column=7, value=item["amount"])
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    # ── Current disbursements ─────────────────────────────────────────────────
    curr_wit_hdr = prev_wit_data_end + 2
    wc(curr_wit_hdr, 2, "Un-reconciled  Disbursement:", bold=True)
    ws.cell(row=curr_wit_hdr, column=4, value=p_label)
    curr_wit_data_start = curr_wit_hdr + 2
    curr_wit_data_end   = curr_wit_data_start - 1

    for item in curr_wit_groups:
        row = curr_wit_data_end + 1
        curr_wit_data_end = row
        ws.cell(row=row, column=2, value=item["date"])
        if isinstance(item["date"], datetime):
            ws.cell(row=row, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=row, column=3, value=item["type"])
        ws.cell(row=row, column=4, value=item["d_key"])
        ws.cell(row=row, column=5, value=item["desc"])
        c = ws.cell(row=row, column=7, value=item["amount"])
        c.number_format = '#,##0.00;[Red]-#,##0.00'

    total_gl_wit_row = curr_wit_data_end + 2
    wc(total_gl_wit_row, 6, "Total")
    c = ws.cell(row=total_gl_wit_row, column=7, value=total_unrec_gl_wit)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    sub_total_row = total_gl_wit_row + 2
    wc(sub_total_row, 6, "Sub Total")
    c = ws.cell(row=sub_total_row, column=7, value=gl_sub_total)
    c.number_format = '#,##0.00;[Red]-#,##0.00'

    adj_book_row = sub_total_row + 2
    wc(adj_book_row, 6, "Adjusted book balance", bold=True)
    c = ws.cell(row=adj_book_row, column=8, value=adjusted_book)
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    c.font = Font(bold=True)

    var_row = adj_book_row + 2
    wc(var_row, 6, "Variance (s/b zero)", bold=True)
    c = ws.cell(row=var_row, column=8, value=variance)
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    c.font = Font(bold=True, color="FF0000" if abs(variance) > 0.01 else "000000")

    # ── Coding legend ─────────────────────────────────────────────────────────
    legend_row = var_row + 5
    wc(legend_row,   8, "Coding",  bold=True)
    wc(legend_row,   9, "Comments", bold=True)
    wc(legend_row+1, 8,  1)
    wc(legend_row+1, 9, "Un reconcile deposit transactions")
    wc(legend_row+2, 8, -1)
    wc(legend_row+2, 9, "Un reconcile withdrawal transactions")
    wc(legend_row+3, 8,  2)
    wc(legend_row+3, 9, "Bank deposit transactions for the current/reporting month match the GL deposit transactions for the current/reporting month.")
    wc(legend_row+4, 8, -2)
    wc(legend_row+4, 9, "Bank withdrawal transactions for the current/reporting month match the GL withdrawal transactions for the current/reporting month.")
    wc(legend_row+5, 8,  3)
    wc(legend_row+5, 9, "The bank deposit transactions of the previous month match the GL deposit transactions of the current/reporting month; and the GL deposit transactions of the previous month match the bank deposit transactions of the current/reporting month.")
    wc(legend_row+6, 8, -3)
    wc(legend_row+6, 9, "The bank withdrawal transactions of the previous month match the GL withdrawal transactions of the current/reporting month; and the GL withdrawal transactions of the previous month match the bank withdrawal transactions of the current/reporting month.")
    wc(legend_row+7, 8, "4 & -4")
    wc(legend_row+7, 9, "Contra or Reverse Transaction within the Bank or GL (4 for deposit entries and -4 for withdrawal entries)")

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Bank Statement
# ═══════════════════════════════════════════════════════════════════════════════
def build_bank_statement(wb, r):
    ws = wb.create_sheet("Bank Statement")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 80

    hdr_font = Font(bold=True, size=11)
    for col, title in enumerate(["Date", "Amount", "Check", "Code", "Description"], 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    gl_by_id = {g["id"]: g for g in r["all_gl"]}

    sorted_bank = sorted(
        r["all_bank"],
        key=lambda b: (parse_date2(b.get("date")) or datetime.min, b.get("amount", 0) or 0)
    )

    row = 2
    unrec_dep_total  = 0.0
    unrec_with_total = 0.0

    for b in sorted_bank:
        code = bank_code(b, gl_by_id)
        amt  = b.get("amount", 0) or 0
        chk  = b.get("check_number") or ""
        desc = b.get("description", "") or ""
        date_val = b.get("date") or ""

        ws.cell(row=row, column=1, value=date_val)
        c = ws.cell(row=row, column=2, value=amt)
        c.number_format = '#,##0.00;[Red]-#,##0.00'
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=chk if chk else "")
        ws.cell(row=row, column=4, value=code)
        ws.cell(row=row, column=5, value=desc)

        if code == 1:
            unrec_dep_total += amt
        elif code == -1:
            unrec_with_total += amt

        row += 1

    # Summary rows
    row += 2
    ws.cell(row=row, column=1, value=1)
    c = ws.cell(row=row, column=2, value=round(unrec_dep_total, 2))
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=3, value="Un-Reconciled Deposits")

    row += 2
    ws.cell(row=row, column=1, value=-1)
    c = ws.cell(row=row, column=2, value=round(unrec_with_total, 2))
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=row, column=3, value="Un-Reconciled Withdrawals")

    ws.freeze_panes = "A2"
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 – GL
# ═══════════════════════════════════════════════════════════════════════════════
def build_gl(wb, r, prop):
    ws = wb.create_sheet("GL")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGHIJKL", [18, 13, 12, 16, 36, 36, 16, 14, 14, 14, 16, 8]):
        ws.column_dimensions[col].width = w

    all_gl    = r["all_gl"]
    bank_by_id = {b["id"]: b for b in r["all_bank"]}

    # Derive date range and period info from all_gl + bank
    gl_dates  = [parse_date2(g.get("date")) for g in all_gl if g.get("date")]
    gl_dates  = [d for d in gl_dates if d]
    bank_dates = [parse_date2(b.get("date")) for b in r["all_bank"] if b.get("date")]
    bank_dates = [d for d in bank_dates if d]
    all_dates  = gl_dates + bank_dates
    if all_dates:
        max_d = max(all_dates)
    else:
        max_d = datetime.today()
    period_start = datetime(max_d.year, max_d.month, 1)
    period_end   = datetime(max_d.year, max_d.month,
                            calendar.monthrange(max_d.year, max_d.month)[1])
    date_range_str = (
        f"{period_start.strftime('%m/%d/%y')} - {period_end.strftime('%m/%d/%y')}"
        f"  (accrual basis)"
    )

    # ── Header rows ───────────────────────────────────────────────────────────
    ws.cell(row=1, column=2, value="General Ledger").font = Font(bold=True, size=12)
    ws.cell(row=2, column=2, value=prop.prop_display)
    ws.cell(row=3, column=2, value=f"Detail Date Range: {date_range_str}")
    ws.cell(row=4, column=1, value=prop.gl_account_id)
    gl_opening = r.get("gl_opening_balance", 0.0) or 0.0
    c = ws.cell(row=4, column=10, value=gl_opening)
    c.number_format = '#,##0.00'

    # ── Column headers ────────────────────────────────────────────────────────
    hdr_labels = ["Property Name", "Date", "Control", "Reference", "Description",
                  "Remarks", "Deposit Number", "Debit", "Credit", "Balance", "Reconcile Date ", "Code"]
    hdr_font = Font(bold=True, size=10)
    for col, lbl in enumerate(hdr_labels, 1):
        c = ws.cell(row=5, column=col, value=lbl)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    sorted_gl = sorted(all_gl, key=lambda g: (parse_date2(g.get("date")) or datetime.min, g["id"]))

    running_balance = gl_opening
    total_dr = 0.0
    total_cr = 0.0
    data_row = 6

    for g in sorted_gl:
        dr  = g.get("debit",  0) or 0.0
        cr  = g.get("credit", 0) or 0.0
        running_balance = round(running_balance + dr - cr, 2)
        total_dr += dr
        total_cr += cr

        code       = gl_code(g, bank_by_id)
        rec_date   = gl_reconcile_date(g, bank_by_id)
        date_val   = g.get("date") or ""

        ws.cell(row=data_row, column=1, value=g.get("property", ""))
        ws.cell(row=data_row, column=2, value=date_val)
        ws.cell(row=data_row, column=3, value=g.get("control", ""))
        ws.cell(row=data_row, column=4, value=g.get("ref", ""))
        ws.cell(row=data_row, column=5, value=g.get("desc", ""))
        ws.cell(row=data_row, column=6, value=g.get("remarks", ""))
        ws.cell(row=data_row, column=7, value=g.get("deposit_num", ""))

        if dr:
            c = ws.cell(row=data_row, column=8, value=dr)
            c.number_format = '#,##0.00'
        if cr:
            c = ws.cell(row=data_row, column=9, value=cr)
            c.number_format = '#,##0.00'

        c = ws.cell(row=data_row, column=10, value=running_balance)
        c.number_format = '#,##0.00'

        if rec_date:
            ws.cell(row=data_row, column=11, value=rec_date)
        ws.cell(row=data_row, column=12, value=code)

        data_row += 1

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = data_row
    ws.cell(row=total_row, column=6, value="Total:").font = Font(bold=True)
    c = ws.cell(row=total_row, column=8, value=round(total_dr, 2))
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)
    c = ws.cell(row=total_row, column=9, value=round(total_cr, 2))
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)
    c = ws.cell(row=total_row, column=10, value=running_balance)
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)

    # ── Un-reconciled summary rows ────────────────────────────────────────────
    unrec_dep_gl  = round(sum((g.get("debit")  or 0) for g in all_gl if not g.get("matched") and (g.get("debit")  or 0) > 0), 2)
    unrec_with_gl = round(sum((g.get("credit") or 0) for g in all_gl if not g.get("matched") and (g.get("credit") or 0) > 0), 2)

    sum_row_dep  = total_row + 3
    sum_row_with = total_row + 5

    ws.cell(row=sum_row_dep, column=6, value=1)
    c = ws.cell(row=sum_row_dep, column=8, value=unrec_dep_gl)
    c.number_format = '#,##0.00'
    ws.cell(row=sum_row_dep, column=9, value="Un- Reconciled Deposits")

    ws.cell(row=sum_row_with, column=6, value=-1)
    c = ws.cell(row=sum_row_with, column=8, value=-unrec_with_gl)
    c.number_format = '#,##0.00;[Red]-#,##0.00'
    ws.cell(row=sum_row_with, column=9, value="Un-Reconciled Withdrawals")

    ws.freeze_panes = "A6"
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 – Un-Reconcile transactions
# ═══════════════════════════════════════════════════════════════════════════════
def build_unrec_transactions(wb, r, prop):
    ws = wb.create_sheet("Un-Reconcile transactions")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGHI", [28, 14, 12, 24, 36, 36, 16, 8, 50]):
        ws.column_dimensions[col].width = w

    sc_ids  = set(r.get("stagecoach_ids", []))
    all_bank = r["all_bank"]
    all_gl   = r["all_gl"]

    unrec_bank = [b for b in all_bank
                  if not b.get("match_rule") and b["id"] not in sc_ids]
    unrec_gl   = [g for g in all_gl if not g.get("matched")]

    hdr_font  = Font(bold=True, size=11)
    col_hdrs  = ["Property Name", "Date", "Control", "Reference",
                 "Description", "Remarks", "Amount", "Comment"]

    def write_section_header(row, label):
        ws.cell(row=row, column=1, value=label).font = hdr_font

    def write_col_headers(row):
        for col, lbl in enumerate(col_hdrs, 1):
            c = ws.cell(row=row, column=col, value=lbl)
            c.font = Font(bold=True)
            c.border = _BORDER

    def write_total(row, total):
        c = ws.cell(row=row, column=7, value=round(total, 2))
        c.number_format = '#,##0.00;[Red]-#,##0.00'
        c.font = Font(bold=True)

    # ── Bank section ──────────────────────────────────────────────────────────
    write_section_header(1, "Bank")
    write_col_headers(2)

    bank_total = 0.0
    row = 3
    for b in sorted(unrec_bank, key=lambda x: parse_date2(x.get("date")) or datetime.min):
        amt    = b.get("amount", 0) or 0
        ctype  = "AR" if amt >= 0 else "AP"
        ws.cell(row=row, column=1, value=prop.prop_label_bank)
        ws.cell(row=row, column=2, value=b.get("date") or "")
        ws.cell(row=row, column=3, value=None)
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=None)
        ws.cell(row=row, column=6, value=b.get("description", ""))
        c = ws.cell(row=row, column=7, value=amt)
        c.number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=8, value=ctype)
        bank_total += amt
        row += 1

    bank_total_row = row
    write_total(bank_total_row, bank_total)
    row += 2

    # ── GL section ────────────────────────────────────────────────────────────
    write_section_header(row, "GL")
    row += 1
    write_col_headers(row)
    row += 1

    # Sort: AR (deposits) first, then AP (withdrawals), each sorted by date then control
    ar_gl   = sorted([g for g in unrec_gl if (g.get("debit")  or 0) > 0],
                     key=lambda g: (parse_date2(g.get("date")) or datetime.min, g.get("control","") or ""))
    ap_gl   = sorted([g for g in unrec_gl if (g.get("credit") or 0) > 0],
                     key=lambda g: (parse_date2(g.get("date")) or datetime.min, g.get("control","") or ""))

    ar_total = 0.0
    ap_total = 0.0

    for g in ar_gl:
        amt = g.get("debit", 0) or 0
        ws.cell(row=row, column=1, value=g.get("property", ""))
        ws.cell(row=row, column=2, value=g.get("date") or "")
        ws.cell(row=row, column=3, value=g.get("control", ""))
        ws.cell(row=row, column=4, value=g.get("ref", ""))
        ws.cell(row=row, column=5, value=g.get("desc", ""))
        ws.cell(row=row, column=6, value=g.get("remarks", ""))
        c = ws.cell(row=row, column=7, value=amt)
        c.number_format = '#,##0.00'
        ws.cell(row=row, column=8, value="AR")
        ar_total += amt
        row += 1

    # AR subtotal
    write_total(row, ar_total)
    row += 1

    for g in ap_gl:
        amt = -(g.get("credit", 0) or 0)   # show as negative
        ws.cell(row=row, column=1, value=g.get("property", ""))
        ws.cell(row=row, column=2, value=g.get("date") or "")
        ws.cell(row=row, column=3, value=g.get("control", ""))
        ws.cell(row=row, column=4, value=g.get("ref", ""))
        ws.cell(row=row, column=5, value=g.get("desc", ""))
        ws.cell(row=row, column=6, value=g.get("remarks", ""))
        c = ws.cell(row=row, column=7, value=amt)
        c.number_format = '#,##0.00;[Red]-#,##0.00'
        ws.cell(row=row, column=8, value="AP")
        ap_total += amt
        row += 1

    # AP subtotal
    write_total(row, ap_total)

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════════════════
def run(results: dict, output_path: str, prop) -> str:
    """Write the 4-sheet output workbook. `results` is the reconcile_ph dict
    (JSON-serialized form: dates as MM/DD/YYYY strings)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary(wb, results, prop)
    build_bank_statement(wb, results)
    build_gl(wb, results, prop)
    build_unrec_transactions(wb, results, prop)

    wb.save(output_path)
    print(f"Saved → {output_path}")
    return output_path
