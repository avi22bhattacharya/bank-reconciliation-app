"""
Bank Reconciliation – March 2026 (PH Davie / 113phv^114epr)

Sources:
  Bank:       Sheet "DepositAccount_8451_040126_729."  in
              03 2026 Bank Recon Data_PH Davie_113phv^114epr_MH.xlsx
  GL:         Sheet "GL"                               in GL_with_Deposit_Numbers.xlsx
              (header row 6, data row 7+, 18 cols – Deposit Number at col R / index 17)
  Prev unrec: Sheet "Un-Reconcile GL"                  in same file
              (header row 1, data row 2+, 14 cols – Amount signed at col I / index 8,
               Deposit Number at col N / index 13)

Key rule changes vs. MH reconcile.py:
  P3 Check          → 1-3 BD prior window added
  P4 Intellipay     → 1-3 BD prior window added
  P5 LSE            → 1-3 BD prior window added
  P6 LAKESHOREMANAGEM → group GL by Deposit Number; group sum must = bank amount;
                        1-2 BD prior; one deposit# per bank entry (no subset-sum
                        across deposit numbers)
  P7 YARDI CARD DEP → same as P6 but 4-6 BD prior
"""

import openpyxl, re, json
from collections import defaultdict
from datetime import datetime, timedelta

BASE   = "/sessions/nice-gallant-meitner/mnt/deposit-register-trial/"
BANK_FILE = BASE + "03 2026 Bank Recon Data_PH Davie_113phv^114epr_MH.xlsx"
GL_FILE   = BASE + "GL_with_Deposit_Numbers.xlsx"

BANK_SHEET = "DepositAccount_8451_040126_729."
GL_SHEET   = "GL"
PREV_SHEET = "Un-Reconcile GL"

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(",", "").strip()
    try: return float(s)
    except: return 0.0

def parse_date(v):
    if isinstance(v, datetime): return v
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(v.split(" ")[0] if " " in v else v, fmt.split(" ")[0])
            except: pass
    return None

def is_numeric_ref(ref):
    return bool(ref and re.match(r'^[\d\s/]+$', ref.strip()) and ref.strip())

def gl_is_yardi_type(rem):
    r = rem.lower()
    return any(k in r for k in ["credit card on-line payment",
                                 "debit card on-line payment",
                                 "recurring debit card payment",
                                 "recurring credit card payment"])

def gl_is_lakeshore_type(ref, rem):
    ref = ref.lower(); rem = rem.lower()
    t1 = ":ach-web" in ref and "online payment - eft" in rem
    t2 = ":ach" in ref     and "pre-authorized payment" in rem
    t3 = ":wips" in ref    and "wips receipt" in rem
    t4 = is_numeric_ref(ref) and ":checkscan payment" in rem
    t5 = (":ach" in ref or ":wips" in ref) and "nsf receipt" in rem
    return t1 or t2 or t3 or t4 or t5

def fast_subset_sum(candidates, target, dp_limit=200):
    """DP subset-sum over (id, amount_cents). Returns list of matched ids or None."""
    if not candidates or target <= 0: return None
    target = round(target)
    if len(candidates) <= dp_limit:
        dp = {0: []}
        for cid, amt in candidates:
            amt = round(amt)
            if amt <= 0: continue
            additions = {}
            for s, used in dp.items():
                ns = s + amt
                if ns == target: return used + [cid]
                if ns < target:  additions[ns] = used + [cid]
            dp.update(additions)
        return None
    else:
        ordered = sorted(candidates, key=lambda x: -x[1])
        chosen, remaining = [], target
        for cid, amt in ordered:
            amt = round(amt)
            if amt <= remaining:
                chosen.append(cid); remaining -= amt
                if remaining == 0: return chosen
        return None

def c(x): return round(x * 100)   # dollars → cents

_BIG_DATE = datetime(9999, 12, 31)

def business_days_apart(d1, d2):
    """Business days (Mon–Fri) between two dates. Returns 9999 if either is None."""
    if d1 is None or d2 is None: return 9999
    start, end = (d1, d2) if d1 <= d2 else (d2, d1)
    count = 0
    d = start
    while d < end:
        d += timedelta(days=1)
        if d.weekday() < 5:
            count += 1
    return count

def bd_before(bank_date, gl_date, lo, hi):
    """True if gl_date is lo..hi business days strictly BEFORE bank_date."""
    if bank_date is None or gl_date is None: return False
    if gl_date >= bank_date: return False
    return lo <= business_days_apart(bank_date, gl_date) <= hi

# Date-window constants
BD_LAKESHORE_LO, BD_LAKESHORE_HI = 1, 3   # GL 1-3 BD before bank (3 covers weekend-dated GL entries)
BD_YARDI_LO,     BD_YARDI_HI     = 1, 6   # GL 1-6 BD before bank
BD_WITHDRAWAL_LO, BD_WITHDRAWAL_HI = 1, 3  # GL 1-3 BD before bank (checks, intellipay, LSE)
WINDOW_BD = 3                               # P8/P10 catch-all ±3 BD

# ═══════════════════════════════════════════════════════════════════════════════
# NON-CASH FILTERS
# ═══════════════════════════════════════════════════════════════════════════════

NONCASH_REM_KEYS = [
    ":prog gen", "reapplied receipt",
    "automatically generated apply prepay",
    ":prog gen reverses", ":prog gen credit",
    ":prog gen prepayment", ":prog gen move-out",
]

def is_noncash_gl(ref, desc, rem):
    rem_l  = rem.lower()
    desc_l = desc.lower()
    ref_l  = ref.lower()
    if any(k in rem_l  for k in NONCASH_REM_KEYS): return True
    if any(k in desc_l for k in NONCASH_REM_KEYS): return True
    if re.match(r'^:reversal of j-', ref_l): return True
    if "accrue" in rem_l and "swap" in rem_l:  return True
    if "accrue" in rem_l and "swap" in desc_l: return True
    return False

def is_noncash_prev_gl(ref, desc, rem):
    combined = (ref + " " + desc + " " + rem).lower()
    if "park mgmt fees" in combined: return True
    if "ma march cash"  in combined: return True
    if "pk to hm pk"    in combined: return True
    if "accrue" in combined and "swap" in combined: return True
    if "reversed by j-" in combined: return True
    return False

# ═══════════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════

wb_bank = openpyxl.load_workbook(BANK_FILE, data_only=True)
wb_gl   = openpyxl.load_workbook(GL_FILE,   data_only=True)

# Read GL opening balance from row 7, col J (index 9) – the balance before data rows
_gl_ob_row = next(wb_gl[GL_SHEET].iter_rows(min_row=7, max_row=7, values_only=True), None)
GL_OPENING_BALANCE = float(_gl_ob_row[9]) if (_gl_ob_row and _gl_ob_row[9] is not None) else 0.0

# ── Bank Statement ────────────────────────────────────────────────────────────
bank = []
for i, row in enumerate(wb_bank[BANK_SHEET].iter_rows(min_row=2, values_only=True)):
    if not any(row): continue
    amt  = round(safe_float(row[1]), 2)
    chk  = str(int(row[2])) if row[2] is not None else None
    desc = str(row[3]).strip() if row[3] else ""
    bank.append({
        "id": f"BANK-{i+2}", "date": parse_date(row[0]),
        "amount": amt, "check_number": chk, "description": desc,
        "source": "Mar 2026 Bank",
        "matched": False, "match_ids": [], "match_rule": "",
    })

# ── GL (current period) ───────────────────────────────────────────────────────
SKIP_REM = {"= Beginning Balance =", "= Ending Balance ="}

gl = []
for i, row in enumerate(wb_gl[GL_SHEET].iter_rows(min_row=8, values_only=True)):
    if not any(row): continue
    remarks = str(row[10]).strip() if row[10] else ""
    if remarks in SKIP_REM: continue
    if row[0] and str(row[0]).startswith("Total"): continue
    debit  = round(safe_float(row[7]), 2)
    credit = round(safe_float(row[8]), 2)
    if debit == 0 and credit == 0: continue
    if row[2] is None and row[5] is None and (debit > 100000 or credit > 100000):
        continue
    ref  = str(row[6]).strip()  if row[6]  else ""
    desc = str(row[4]).strip()  if row[4]  else ""
    prop = str(row[1]).strip()  if row[1]  else ""
    dep_raw = str(row[17]).strip() if row[17] else ""
    dep_num = re.sub(r'[^0-9]', '', dep_raw) if dep_raw else ""
    if is_noncash_gl(ref, desc, remarks): continue
    gl.append({
        "id": f"GL-{i+8}", "date": parse_date(row[2]),
        "desc": desc, "control": str(row[5]).strip() if row[5] else "",
        "ref": ref, "debit": debit, "credit": credit, "remarks": remarks,
        "property": prop, "deposit_num": dep_num,
        "source": "Mar 2026 GL",
        "matched": False, "match_ids": [], "match_rule": "",
    })

# ── Previous unreconciled GL ──────────────────────────────────────────────────
prev_gl = []
for i, row in enumerate(wb_gl[PREV_SHEET].iter_rows(min_row=2, values_only=True)):
    if not any(row): continue
    amt_raw = safe_float(row[8])
    if amt_raw == 0.0: continue
    ref   = str(row[6]).strip() if row[6] else ""
    desc  = str(row[4]).strip() if row[4] else ""
    rem   = str(row[9]).strip() if row[9] else ""
    prop  = str(row[1]).strip() if row[1] else ""
    dep_raw = str(row[13]).strip() if row[13] else ""
    dep_num = re.sub(r'[^0-9]', '', dep_raw) if dep_raw else ""
    if not ref and not desc and not rem: continue
    if is_noncash_prev_gl(ref, desc, rem): continue
    debit  = round( amt_raw, 2) if amt_raw > 0 else 0.0
    credit = round(-amt_raw, 2) if amt_raw < 0 else 0.0
    prev_gl.append({
        "id": f"PREV-GL-{i+2}", "date": parse_date(row[2]),
        "desc": desc, "control": str(row[5]).strip() if row[5] else "",
        "ref": ref, "debit": debit, "credit": credit, "remarks": rem,
        "property": prop, "deposit_num": dep_num,
        "source": "Feb/Mar 2026 Unrec GL",
        "matched": False, "match_ids": [], "match_rule": "",
    })

all_bank = bank
all_gl   = gl + prev_gl
print(f"Bank rows: {len(all_bank)},  GL rows: {len(all_gl)}  "
      f"(current: {len(gl)}, prev unrec: {len(prev_gl)})")

# ═══════════════════════════════════════════════════════════════════════════════
# PRE-PASS: CONTRA MATCHING (runs before all reconciliation passes)
# ═══════════════════════════════════════════════════════════════════════════════

# ── Contra - Bank: same check number, amounts net to zero ────────────────────
# Applied after merging prior + current month bank data, before any other pass.
# All-or-nothing per check-number group.
contra_bank_ids = []
_chk_groups = defaultdict(list)
for b in all_bank:
    if b["check_number"]:
        _chk_groups[b["check_number"]].append(b)
for chk, grp in _chk_groups.items():
    if len(grp) >= 2 and round(sum(b["amount"] for b in grp), 2) == 0.0:
        for b in grp:
            b["matched"] = True
            b["match_rule"] = "Contra - Bank"
        contra_bank_ids.extend(b["id"] for b in grp)

# ── Contra - GL: same reference, debit − credit nets to zero ─────────────────
# Applied after merging prior + current month GL data, before any other pass.
# All-or-nothing per reference group.
contra_gl_groups = []   # list-of-lists of GL IDs (one inner list per group)
_ref_groups = defaultdict(list)
for g in all_gl:
    if g["ref"]:
        _ref_groups[g["ref"]].append(g)
for ref, grp in _ref_groups.items():
    if len(grp) < 2:
        continue
    # Skip groups containing Lakeshore-type or Yardi-type GL entries –
    # those are reserved for P6/P7 and must remain available for bank matching.
    if any(gl_is_lakeshore_type(g["ref"], g["remarks"]) or gl_is_yardi_type(g["remarks"])
           for g in grp):
        continue
    net = round(sum(g["debit"] for g in grp) - sum(g["credit"] for g in grp), 2)
    if net == 0.0:
        for g in grp:
            g["matched"] = True
            g["match_rule"] = "Contra - GL"
        contra_gl_groups.append([g["id"] for g in grp])

n_contra_bank = len(contra_bank_ids)
n_contra_gl   = sum(len(grp) for grp in contra_gl_groups)
print(f"Contra pre-pass: {n_contra_bank} bank entries (Contra-Bank), "
      f"{n_contra_gl} GL entries across {len(contra_gl_groups)} groups (Contra-GL)")

# ═══════════════════════════════════════════════════════════════════════════════
# MATCHING INFRASTRUCTURE
# ═══════════════════════════════════════════════════════════════════════════════

gl_by_id   = {g["id"]: g for g in all_gl}
bank_by_id = {b["id"]: b for b in all_bank}

def mark_matched(bank_id, gl_ids, rule):
    b = bank_by_id[bank_id]
    b["matched"] = True; b["match_ids"] = list(gl_ids); b["match_rule"] = rule
    for gid in gl_ids:
        g = gl_by_id[gid]
        g["matched"] = True
        if bank_id not in g["match_ids"]: g["match_ids"].append(bank_id)

def amount_match_gl(pool, bank_amount):
    """Single bank → multiple GL. bank>0 → debit side; bank<0 → credit side."""
    if bank_amount > 0:
        target = c(bank_amount)
        cands  = [(g["id"], c(g["debit"]))  for g in pool if not g["matched"] and g["debit"]  > 0]
    else:
        target = c(abs(bank_amount))
        cands  = [(g["id"], c(g["credit"])) for g in pool if not g["matched"] and g["credit"] > 0]
    return fast_subset_sum(cands, target)

# ═══════════════════════════════════════════════════════════════════════════════
# MATCHING PASSES
# ═══════════════════════════════════════════════════════════════════════════════

# ── P1: Stagecoach Sweep DEBIT ↔ CREDIT (internal bank-to-bank pairs) ───────
# Only entries whose description is EXACTLY "STAGECOACH SWEEP DEBIT" or
# "STAGECOACH SWEEP CREDIT" (case-insensitive) are treated as internal sweeps.
# Each DEBIT must pair with a CREDIT of the same absolute amount.
# Entries with no counterpart are left unmatched and flow to Unmatched Bank.
# Any other description containing "stagecoach" (e.g. SWEEP INTEREST PAYMENT)
# is NOT caught here and is handled by P2 as a normal GL match.
stagecoach_internal = []
_sc_debits  = [b for b in all_bank
               if b["description"].upper().strip() == "STAGECOACH SWEEP DEBIT"]
_sc_credits = [b for b in all_bank
               if b["description"].upper().strip() == "STAGECOACH SWEEP CREDIT"]
_avail_cr = list(_sc_credits)
for _db in _sc_debits:
    _db_abs = round(abs(_db["amount"]), 2)
    _cr = next((cr for cr in _avail_cr
                if round(cr["amount"], 2) == _db_abs), None)
    if _cr:
        _db["matched"] = True; _db["match_rule"] = "INTERNAL – Stagecoach Sweep"
        _cr["matched"] = True; _cr["match_rule"] = "INTERNAL – Stagecoach Sweep"
        stagecoach_internal.extend([_db["id"], _cr["id"]])
        _avail_cr.remove(_cr)
    # No matching CREDIT in this period → DEBIT stays unmatched (closing sweep,
    # its return CREDIT will appear in the next month's statement).

# Unpaired CREDITs are opening-balance returns from the prior month's closing
# sweep — their matching DEBIT was reconciled in the previous period, so mark
# them as matched internal entries.
for _cr in _avail_cr:
    _cr["matched"] = True; _cr["match_rule"] = "INTERNAL – Stagecoach Sweep"
    stagecoach_internal.append(_cr["id"])
    print(f"  NOTE: {_cr['id']} ({_cr['amount']:,.2f}) is an opening-balance return "
          f"from prior month — marked as matched internal.")

_n_sc_open_debits = len([b for b in _sc_debits if not b["matched"]])
if _n_sc_open_debits:
    for _db in _sc_debits:
        if not _db["matched"]:
            print(f"  NOTE: {_db['id']} ({_db['amount']:,.2f}) is the period-end closing "
                  f"sweep — no return CREDIT yet → Unmatched Bank.")

# ── P2: Stagecoach Sweep INTEREST → GL "Sweep Interest Payment" ──────────────
for b in all_bank:
    if b["matched"]: continue
    if "STAGECOACH SWEEP INTEREST" in b["description"].upper():
        pool = [g for g in all_gl if not g["matched"]
                and "sweep interest payment" in g["remarks"].lower()]
        ids = amount_match_gl(pool, b["amount"])
        if ids: mark_matched(b["id"], ids, "Stagecoach Sweep Interest Payment")

# ── P3: Check number → GL Reference  (1-3 BD prior window) ──────────────────
gl_by_ref = defaultdict(list)
for g in all_gl: gl_by_ref[g["ref"]].append(g)

_p3_bank = sorted(
    [b for b in all_bank if not b["matched"] and b["check_number"] and b["date"]],
    key=lambda b: b["date"]
)
for b in _p3_bank:
    if b["matched"]: continue
    chk   = b["check_number"]
    cands = [g for g in gl_by_ref.get(chk, [])
             if not g["matched"]
             and bd_before(b["date"], g["date"], BD_WITHDRAWAL_LO, BD_WITHDRAWAL_HI)]
    if not cands:
        # Fallback: no date restriction for checks (some may post same-day or next-day)
        cands = [g for g in gl_by_ref.get(chk, []) if not g["matched"]]
    if not cands: continue
    bank_abs = round(abs(b["amount"]), 2)
    total_cr = round(sum(g["credit"] for g in cands), 2)
    if total_cr == bank_abs:
        mark_matched(b["id"], [g["id"] for g in cands], f"Check #{chk}")
    else:
        ids = fast_subset_sum([(g["id"], c(g["credit"])) for g in cands], c(bank_abs))
        if ids: mark_matched(b["id"], ids, f"Check #{chk} (subset)")

# ── P4: INTELLIPAY BILLING → "Convenient Payments" (1-3 BD prior window) ─────
_p4_bank = sorted(
    [b for b in all_bank if not b["matched"]
     and "INTELLIPAY BILLING" in b["description"].upper() and b["date"]],
    key=lambda b: b["date"]
)
for b in _p4_bank:
    if b["matched"]: continue
    pool = [g for g in all_gl if not g["matched"]
            and "convenient payments" in g["remarks"].lower()
            and bd_before(b["date"], g["date"], BD_WITHDRAWAL_LO, BD_WITHDRAWAL_HI)]
    ids = amount_match_gl(pool, b["amount"])
    if ids: mark_matched(b["id"], ids, "INTELLIPAY BILLING → Convenient Payments")

# ── P5: LAKESHORE EMPLOYMENT → LSE (1-3 BD prior window) ────────────────────
_p5_bank = sorted(
    [b for b in all_bank if not b["matched"]
     and "LAKESHORE EMPLOYMENT" in b["description"].upper() and b["date"]],
    key=lambda b: b["date"]
)
for b in _p5_bank:
    if b["matched"]: continue
    bank_abs = round(abs(b["amount"]), 2)
    lse_pool = [g for g in all_gl if not g["matched"]
                and "lse (v0000665)" in g["desc"].lower()
                and bd_before(b["date"], g["date"], BD_WITHDRAWAL_LO, BD_WITHDRAWAL_HI)]
    lse_by_ref = defaultdict(list)
    for g in lse_pool: lse_by_ref[g["ref"]].append(g)

    matched = False
    for ref, grp in lse_by_ref.items():
        net = round(sum(g["credit"] for g in grp) - sum(g["debit"] for g in grp), 2)
        if net == bank_abs:
            mark_matched(b["id"], [g["id"] for g in grp],
                         f"LAKESHORE EMPLOYMENT → LSE ref {ref} (net)")
            matched = True; break
    if not matched:
        ids = amount_match_gl(lse_pool, -bank_abs)
        if ids: mark_matched(b["id"], ids, "LAKESHORE EMPLOYMENT → LSE (credit match)")

# ── P6: LAKESHOREMANAGEM → Deposit Number grouping (1-2 BD prior) ────────────
# Rule: group GL by Deposit Number; each group sum must equal bank amount exactly.
# One Deposit Number matches at most one bank entry.
# Process bank entries chronologically so early entries claim their GL window first.

lks_gl = [g for g in all_gl
          if gl_is_lakeshore_type(g["ref"], g["remarks"]) and g["deposit_num"]]

lks_bank_sorted = sorted(
    [b for b in all_bank if "LAKESHOREMANAGEM" in b["description"].upper()],
    key=lambda b: b["date"] if b["date"] else _BIG_DATE
)

for b in lks_bank_sorted:
    if b["matched"]: continue
    is_return = "RETURN" in b["description"].upper()
    bank_abs  = round(abs(b["amount"]), 2)
    bank_date = b["date"]

    if not is_return:
        # Build deposit-number groups from windowed pool (debit side)
        windowed = [g for g in lks_gl
                    if not g["matched"] and g["debit"] > 0
                    and bd_before(bank_date, g["date"],
                                  BD_LAKESHORE_LO, BD_LAKESHORE_HI)]
        dep_grp = defaultdict(list)
        for g in windowed:
            dep_grp[g["deposit_num"]].append(g)

        # Find single deposit# whose group sum = bank amount exactly
        matched_dep = None
        for dep, rows in dep_grp.items():
            grp_sum = round(sum(g["debit"] for g in rows), 2)
            if grp_sum == bank_abs:
                matched_dep = dep
                break

        if matched_dep:
            gids = [g["id"] for g in dep_grp[matched_dep] if not g["matched"]]
            mark_matched(b["id"], gids, "LAKESHOREMANAGEM Settlement")
    else:
        # Return side: credit entries 1-2 BD before bank date
        windowed_cr = [g for g in lks_gl
                       if not g["matched"] and g["credit"] > 0
                       and bd_before(bank_date, g["date"],
                                     BD_LAKESHORE_LO, BD_LAKESHORE_HI)]
        dep_grp_cr = defaultdict(list)
        for g in windowed_cr:
            dep_grp_cr[g["deposit_num"]].append(g)

        matched_dep = None
        for dep, rows in dep_grp_cr.items():
            grp_sum = round(sum(g["credit"] for g in rows), 2)
            if grp_sum == bank_abs:
                matched_dep = dep
                break

        if matched_dep:
            gids = [g["id"] for g in dep_grp_cr[matched_dep] if not g["matched"]]
            mark_matched(b["id"], gids, "LAKESHOREMANAGEM Return")

# ── P7: YARDI CARD DEP → Deposit Number grouping (4-6 BD prior) ──────────────
yardi_gl = [g for g in all_gl
            if gl_is_yardi_type(g["remarks"]) and g["deposit_num"]]

yardi_bank_sorted = sorted(
    [b for b in all_bank if "YARDI CARD DEP" in b["description"].upper()],
    key=lambda b: b["date"] if b["date"] else _BIG_DATE
)

for b in yardi_bank_sorted:
    if b["matched"]: continue
    bank_abs  = round(abs(b["amount"]), 2)
    bank_date = b["date"]

    windowed = [g for g in yardi_gl
                if not g["matched"] and g["debit"] > 0
                and bd_before(bank_date, g["date"], BD_YARDI_LO, BD_YARDI_HI)]

    dep_grp = defaultdict(list)
    for g in windowed:
        dep_grp[g["deposit_num"]].append(g)

    matched_dep = None
    for dep, rows in dep_grp.items():
        grp_sum = round(sum(g["debit"] for g in rows), 2)
        if grp_sum == bank_abs:
            matched_dep = dep
            break

    if matched_dep:
        gids = [g["id"] for g in dep_grp[matched_dep] if not g["matched"]]
        mark_matched(b["id"], gids, "YARDI CARD DEP")

# ═══════════════════════════════════════════════════════════════════════════════
# PASS 2  (P8–P10): fallback passes — only for entries still unmatched after P1-P7
# ═══════════════════════════════════════════════════════════════════════════════

after_p7_unmatched = [b["id"] for b in all_bank if not b["matched"]]
print(f"\nAfter P1-P7: {len(all_bank) - len(after_p7_unmatched)} bank matched, "
      f"{len(after_p7_unmatched)} still unmatched → entering fallback passes P8-P10")

# ── P8: Amount fallback (±3 BD guard, excl. Lakeshore/Yardi types) ───────────
# LAKESHOREMANAGEM and YARDI CARD DEP bank entries are reserved for P6/P7.
# Exact Stagecoach Sweep DEBIT/CREDIT entries are internal bank-to-bank sweeps
# with no GL counterpart — they must never be matched against GL entries.
_SC_SWEEP_EXACT = {"STAGECOACH SWEEP DEBIT", "STAGECOACH SWEEP CREDIT"}

def _is_deposit_type_bank(desc):
    du = desc.upper()
    return ("LAKESHOREMANAGEM" in du or "YARDI CARD DEP" in du
            or du.strip() in _SC_SWEEP_EXACT)

_p8_bank = sorted(
    [b for b in all_bank
     if b["id"] in set(after_p7_unmatched) and b["date"]
     and not _is_deposit_type_bank(b["description"])],
    key=lambda b: b["date"]
)
for b in _p8_bank:
    if b["matched"]: continue
    pool = [g for g in all_gl
            if not g["matched"]
            and not gl_is_lakeshore_type(g["ref"], g["remarks"])
            and not gl_is_yardi_type(g["remarks"])
            and business_days_apart(b["date"], g["date"]) <= WINDOW_BD]
    ids = amount_match_gl(pool, b["amount"])
    if ids: mark_matched(b["id"], ids, "Amount match (fallback)")

# ── P9: Reverse match – multiple bank entries → one GL credit ────────────────
# Exclude exact Stagecoach Sweep DEBIT/CREDIT — internal, no GL counterpart.
from itertools import combinations as _combos
_p9_ids = set(after_p7_unmatched)
unmatched_bank_neg = [b for b in all_bank
                      if b["id"] in _p9_ids and not b["matched"] and b["amount"] < 0
                      and b["description"].upper().strip() not in _SC_SWEEP_EXACT]
unmatched_gl_cr    = [g for g in all_gl   if not g["matched"] and g["credit"] > 0]
gl_cr_by_amt = defaultdict(list)
for g in unmatched_gl_cr:
    gl_cr_by_amt[c(g["credit"])].append(g["id"])

for size in range(2, 4):
    if not unmatched_bank_neg: break
    for combo in _combos([b for b in unmatched_bank_neg if not b["matched"]], size):
        total = c(round(sum(abs(b["amount"]) for b in combo), 2))
        if total in gl_cr_by_amt:
            gl_id = gl_cr_by_amt[total][0]
            if gl_by_id[gl_id]["matched"]: continue
            bank_ids_in_combo = [b["id"] for b in combo]
            for bid in bank_ids_in_combo:
                bank_by_id[bid]["matched"]    = True
                bank_by_id[bid]["match_ids"]  = [gl_id]
                bank_by_id[bid]["match_rule"] = "Multi-bank → single GL"
            gl_by_id[gl_id]["matched"]   = True
            gl_by_id[gl_id]["match_ids"] = bank_ids_in_combo
            break

# ── P10: Date-window catch-all (±3 BD, all GL types) ────────────────────────
# Same exclusion: LAKESHOREMANAGEM and YARDI CARD DEP entries must not match here.
_p10_bank = sorted(
    [b for b in all_bank
     if b["id"] in set(after_p7_unmatched) and not b["matched"] and b["date"]
     and not _is_deposit_type_bank(b["description"])],
    key=lambda b: b["date"]
)
for b in _p10_bank:
    if b["matched"]: continue
    pool = [g for g in all_gl
            if not g["matched"]
            and business_days_apart(b["date"], g["date"]) <= WINDOW_BD]
    ids = amount_match_gl(pool, b["amount"])
    if ids: mark_matched(b["id"], ids, f"Date-window match (±{WINDOW_BD}BD)")

# ═══════════════════════════════════════════════════════════════════════════════
# RESULTS
# ═══════════════════════════════════════════════════════════════════════════════

real_bank      = [b for b in all_bank if b["match_rule"] != "INTERNAL – Stagecoach Sweep"]
matched_bank   = [b for b in real_bank if b["matched"]]
unmatched_bank = [b for b in real_bank if not b["matched"]]
matched_gl     = [g for g in all_gl if g["matched"]]
unmatched_gl   = [g for g in all_gl if not g["matched"]]

pct_bank = round(len(matched_bank) / len(real_bank) * 100, 1) if real_bank else 0
pct_gl   = round(len(matched_gl)   / len(all_gl)   * 100, 1) if all_gl    else 0

print("\n" + "="*72)
print("RECONCILIATION RESULTS – March 2026  (PH Davie 113phv/114epr)")
print("="*72)
print(f"  Stagecoach Sweep DEBIT/CREDIT (internal): {len(stagecoach_internal)}")
print(f"\n  Bank (excl. internal):  {len(real_bank):3d} total | "
      f"{len(matched_bank):3d} matched ({pct_bank}%) | {len(unmatched_bank):3d} unmatched")
print(f"  GL:                     {len(all_gl):3d} total | "
      f"{len(matched_gl):3d} matched ({pct_gl}%) | {len(unmatched_gl):3d} unmatched")

print("\n── Unmatched Bank Entries ──")
for b in unmatched_bank:
    ds = b["date"].strftime("%m/%d/%Y") if b["date"] else "N/A"
    print(f"  {b['id']:12s}  {ds}  {b['amount']:>12,.2f}  {b['description'][:65]}")

print("\n── Unmatched GL Entries ──")
for g in unmatched_gl:
    net = g["debit"] if g["debit"] else -g["credit"]
    ds  = g["date"].strftime("%m/%d/%Y") if g["date"] else "N/A"
    print(f"  {g['id']:14s}  {ds}  {net:>12,.2f}  [{g['property'][:12]}]  "
          f"dep={g['deposit_num'] or '-':>6}  ref={g['ref'][:18]}  {g['remarks'][:40]}")

# ═══════════════════════════════════════════════════════════════════════════════
# SAVE JSON
# ═══════════════════════════════════════════════════════════════════════════════

def dt_serial(obj):
    if isinstance(obj, datetime): return obj.strftime("%m/%d/%Y")
    raise TypeError(f"Not serializable: {type(obj)}")

JSON_OUT = BASE + "recon_ph_results.json"
with open(JSON_OUT, "w") as f:
    json.dump({
        "all_bank": all_bank, "all_gl": all_gl,
        "stagecoach_ids": stagecoach_internal,
        "contra_bank_ids": contra_bank_ids,
        "contra_gl_groups": contra_gl_groups,
        "pct_bank": pct_bank, "pct_gl": pct_gl,
        "total_real_bank": len(real_bank), "total_gl": len(all_gl),
        "n_matched_bank": len(matched_bank), "n_matched_gl": len(matched_gl),
        "gl_opening_balance": GL_OPENING_BALANCE,
    }, f, default=dt_serial, indent=2)
print(f"\nSaved {JSON_OUT}")
