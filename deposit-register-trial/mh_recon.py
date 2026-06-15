"""
MH Reconciliation – Deposit Number Lookup
==========================================
Reads the Deposit Register and one or more GL source sheets, then writes
an Excel file with three new columns appended to each GL sheet:

  • Tenant Code   – extracted from the Person/Description field
  • Unique ID     – Tenant Code + Date + Remarks (the match key)
  • Deposit Number – digits-only deposit number from the Deposit Register

Usage
-----
Edit CONFIG below, then run:
    python mh_recon.py

Requirements
------------
  pip install openpyxl pandas
  LibreOffice must be installed (for .xls → .xlsx conversion)
"""

import os
import re
import subprocess
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG – edit these paths before running
# ──────────────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DEPOSIT_REGISTER = {
    "file":       "Deposit_Register_May_PH davie MH.xlsx",
    "sheet":      "Deposit_Register",
    "header_row": 12,   # 0-based row index of the column-name row
}

GL_SOURCES = [
    {
        "label":       "GL",
        "file":        "03 2026 Bank Recon Data_PH Davie_113phv^114epr_MH.xls",
        "sheet":       "GL",
        "header_row":  5,   # 0-based; rows 0-4 are metadata kept as-is
        "person_col_idx":  4,   # 0-based column index of Person/Description
        "date_col_idx":    2,
        "remarks_col_idx": 10,
    },
    {
        "label":       "Un-Reconcile GL",
        "file":        "02 2026_Open Items_PH Davie_113phv^114epr_MH.xlsx",
        "sheet":       "Un-Reconcile GL",
        "header_row":  0,   # row 0 is the header; no metadata rows above it
        "person_col_idx":  4,
        "date_col_idx":    2,
        "remarks_col_idx": 9,
    },
]

OUTPUT_FILE = "GL_with_Deposit_Numbers.xlsx"

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def resolve(filename):
    return os.path.join(BASE_DIR, filename)


def xls_to_xlsx(path):
    """Convert .xls → .xlsx using LibreOffice and return the new path."""
    tmp = tempfile.mkdtemp()
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp, path],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice failed:\n{result.stderr}")
    base = os.path.splitext(os.path.basename(path))[0] + ".xlsx"
    return os.path.join(tmp, base)


def safe(v):
    """Return a value safe for openpyxl (no NaN, no Series)."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, pd.Series):
        return str(v.iloc[0]) if len(v) > 0 else ""
    return v


def extract_tenant_code(text):
    """'Smith (t0000771)' → 't0000771', else ''."""
    m = re.search(r'\(([tTrR]\d+)\)', str(text))
    return m.group(1).lower() if m else ""


def digits_only(raw):
    """'ACH-RC 37075' → '37075';  'CC-874' → '874';  '1000' → '1000'."""
    d = re.sub(r'[^0-9]', '', str(raw))
    return d if d else str(raw)


def build_lookup(cfg):
    """Build {(tenant_code, date, notes): raw_deposit_number} from Deposit Register."""
    path = resolve(cfg["file"])
    raw  = pd.read_excel(path, sheet_name=cfg["sheet"], header=None)
    raw.columns = raw.iloc[cfg["header_row"]].tolist()
    df = raw.iloc[cfg["header_row"] + 1 :]
    df = df[df["Tenant Code"].notna()].copy()
    df["_tc"]    = df["Tenant Code"].astype(str).str.strip().str.lower()
    df["_date"]  = pd.to_datetime(df["Received Date"], errors="coerce").dt.date
    df["_notes"] = df["Notes"].astype(str).str.strip().str.lower()
    lookup = {}
    for _, row in df.iterrows():
        key = (row["_tc"], row["_date"], row["_notes"])
        if key not in lookup and pd.notna(row["Deposit Number"]):
            lookup[key] = str(row["Deposit Number"]).strip()
    print(f"  Deposit Register loaded: {len(df)} entries, {len(lookup)} unique lookup keys")
    return lookup


def do_lookup(lookup, person, date_val, remarks):
    """Return (tenant_code, unique_id, deposit_number_clean) for one GL row."""
    tc = extract_tenant_code(person)
    if not tc:
        return "", "", ""
    try:
        date = pd.to_datetime(date_val).date()
    except Exception:
        return tc, "", ""
    rem = str(remarks).strip().lower()
    uid = tc + str(date) + rem
    raw = lookup.get((tc, date, rem))
    dep = digits_only(raw) if raw else ""
    return tc, uid, dep


def process_source(cfg, lookup):
    path = resolve(cfg["file"])
    if path.lower().endswith(".xls"):
        path = xls_to_xlsx(path)

    raw = pd.read_excel(path, sheet_name=cfg["sheet"], header=None)
    hr  = cfg["header_row"]
    col_names = [str(v) if pd.notna(v) else "" for v in raw.iloc[hr].tolist()]

    # Deduplicate column names (e.g. two "Description" columns)
    seen, deduped = {}, []
    for c in col_names:
        if c in seen:
            seen[c] += 1
            deduped.append(f"{c} ({seen[c]})")
        else:
            seen[c] = 0
            deduped.append(c)
    col_names = deduped

    meta_rows = [raw.iloc[i].tolist() for i in range(hr)]   # rows above header
    data_rows = []

    p_idx = cfg["person_col_idx"]
    d_idx = cfg["date_col_idx"]
    r_idx = cfg["remarks_col_idx"]

    for i in range(hr + 1, len(raw)):
        row = raw.iloc[i]
        person  = safe(row.iloc[p_idx])
        date_v  = safe(row.iloc[d_idx])
        remarks = safe(row.iloc[r_idx])
        tc, uid, dep = do_lookup(lookup, person, date_v, remarks)
        vals = [safe(row.iloc[c]) for c in range(len(row))] + [tc, uid, dep]
        data_rows.append(vals)

    matched = sum(1 for r in data_rows if r[-1])
    print(f"  [{cfg['label']}] {matched}/{len(data_rows)} rows matched to a Deposit Number")
    return meta_rows, col_names, data_rows


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

BLUE  = PatternFill("solid", fgColor="1F4E79")
BFNT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
YEL   = PatternFill("solid", fgColor="FFD966")
YFNT  = Font(bold=True, color="000000", name="Calibri", size=10)
LYEL  = PatternFill("solid", fgColor="FFFACD")
DFNT  = Font(name="Calibri", size=10)
CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
NEW_COLS = ["Tenant Code", "Unique ID", "Deposit Number"]


def write_gl_sheet(wb, name, meta_rows, col_names, data_rows):
    ws = wb.create_sheet(name)

    for meta in meta_rows:
        ws.append([safe(v) for v in meta])

    full_header = col_names + NEW_COLS
    ws.append(full_header)
    hr = ws.max_row
    for c, h in enumerate(full_header, 1):
        cell = ws.cell(row=hr, column=c)
        if h in NEW_COLS:
            cell.fill = YEL;  cell.font = YFNT
        else:
            cell.fill = BLUE; cell.font = BFNT
        cell.alignment = CTR

    n = len(full_header)
    for row_vals in data_rows:
        ws.append(row_vals)
        dr = ws.max_row
        for c in range(1, n + 1):
            ws.cell(row=dr, column=c).font = DFNT
        ws.cell(row=dr, column=n).fill = LYEL   # highlight Deposit Number

    # Column widths
    for c in range(1, n + 1):
        col_l = get_column_letter(c)
        w = len(str(full_header[c - 1])) + 2
        for r in range(hr, min(hr + 80, ws.max_row + 1)):
            v = ws.cell(row=r, column=c).value
            if v:
                w = max(w, len(str(v)) + 2)
        ws.column_dimensions[col_l].width = min(w, 45)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    print("Loading Deposit Register...")
    lookup = build_lookup(DEPOSIT_REGISTER)

    wb = Workbook()
    wb.remove(wb.active)

    all_stats = []
    for cfg in GL_SOURCES:
        print(f"\nProcessing: {cfg['file']} → sheet '{cfg['sheet']}'")
        meta, cols, data = process_source(cfg, lookup)
        write_gl_sheet(wb, cfg["label"], meta, cols, data)
        matched = sum(1 for r in data if r[-1])
        all_stats.append((cfg["label"], len(data), matched))

    # Summary sheet
    ws_s = wb.create_sheet("Summary")
    ws_s.append(["Sheet", "Total Rows", "Deposit # Matched", "Unmatched"])
    for label, total, matched in all_stats:
        ws_s.append([label, total, matched, total - matched])

    out = resolve(OUTPUT_FILE)
    wb.save(out)
    print(f"\nOutput saved: {out}")


if __name__ == "__main__":
    main()
